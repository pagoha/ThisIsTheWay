#!/usr/bin/env python3
"""
AWS Storage Analyzer
Analyzes storage usage across EC2, RDS, DynamoDB, and AWS Backups in an AWS account.
"""

import boto3
import botocore.session
import csv
import sys
import time
from datetime import datetime
from collections import defaultdict
from botocore.exceptions import ClientError

# Try to import openpyxl for Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will not be available.")
    print("Install with: pip install openpyxl")

class StorageAnalyzer:
    def __init__(self):
        """Initialize the analyzer"""
        self.session = None
        self.regions = []
        self.results = []
        self.account_id = None
        self.account_alias = None
        self.profile_name = None
        self.output_prefix = None
        
    def get_profile_and_account_info(self):
        """Prompt for AWS profile and confirm account details"""
        session = botocore.session.Session()
        profiles = session.available_profiles
        
        if not profiles:
            print("Error: No AWS profiles found. Please configure AWS CLI first.")
            sys.exit(1)
        
        print("\n" + "="*80)
        print("AWS PROFILE SELECTION")
        print("="*80)
        print("\nAvailable AWS profiles:")
        for i, profile in enumerate(profiles, 1):
            print(f"  {i}. {profile}")
        
        # Profile selection logic
        while True:
            try:
                if len(profiles) == 1:
                    print(f"\nOnly one profile available, using '{profiles[0]}'")
                    profile_name = profiles[0]
                    break
                else:
                    selection = input("\nEnter profile number or name (or press Enter for default): ").strip()
                    
                    if not selection:
                        profile_name = "default"
                        break
                    
                    if selection.isdigit() and 1 <= int(selection) <= len(profiles):
                        profile_name = profiles[int(selection) - 1]
                        break
                    elif selection in profiles:
                        profile_name = selection
                        break
                    else:
                        print("Invalid selection. Please try again.")
            except Exception as e:
                print(f"Error in profile selection: {e}")
        
        # Create session and validate
        try:
            self.session = boto3.Session(profile_name=profile_name)
            self.profile_name = profile_name
            sts = self.session.client('sts')
            
            identity = sts.get_caller_identity()
            self.account_id = identity['Account']
            iam_arn = identity['Arn']
            
            # Try to get account alias
            try:
                iam = self.session.client('iam')
                aliases = iam.list_account_aliases()
                if aliases['AccountAliases']:
                    self.account_alias = aliases['AccountAliases'][0]
            except:
                pass
            
            print("\n" + "="*80)
            print("AWS ACCOUNT INFORMATION")
            print("="*80)
            print(f"Profile:        {profile_name}")
            print(f"Account ID:     {self.account_id}")
            if self.account_alias:
                print(f"Account Alias:  {self.account_alias}")
            print(f"IAM ARN:        {iam_arn}")
            print("="*80)
            
            confirmation = input("\nIs this the correct account? (yes/no): ").lower()
            if confirmation not in ['y', 'yes']:
                print("\nAborting operation.")
                sys.exit(0)
            
            return True
            
        except Exception as e:
            print(f"\nError connecting to AWS with profile '{profile_name}': {e}")
            sys.exit(1)
    
    def setup_interactive(self):
        """Interactive setup for the analyzer"""
        print("\n" + "="*80)
        print("AWS STORAGE ANALYZER")
        print("="*80)
        print()
        
        # Get profile and account info first
        self.get_profile_and_account_info()
        
        print()
        
        # Region Selection
        print("="*80)
        print("REGION SELECTION")
        print("="*80)
        print("\nOptions:")
        print("  1. All regions (comprehensive but slower)")
        print("  2. Specific regions (faster)")
        print("  3. Current region only")
        
        choice = input("\nSelect an option [1-3] (default: 2): ").strip() or "2"
        
        if choice == "1":
            print("\nFetching all available regions...")
            self.regions = self._get_all_regions()
            print(f"✓ Will analyze {len(self.regions)} regions")
        elif choice == "3":
            current_region = self.session.region_name or 'us-east-1'
            self.regions = [current_region]
            print(f"✓ Will analyze current region: {current_region}")
        else:  # choice == "2" or default
            print("\nCommon regions:")
            common_regions = [
                'us-east-1', 'us-east-2', 'us-west-1', 'us-west-2',
                'eu-west-1', 'eu-central-1', 'ap-southeast-1', 'ap-northeast-1'
            ]
            for idx, region in enumerate(common_regions, 1):
                print(f"  {idx}. {region}")
            
            print("\nEnter region names separated by commas (e.g., us-east-1,us-west-2)")
            print("Or enter numbers from the list above (e.g., 1,4)")
            region_input = input("Regions: ").strip()
            
            if not region_input:
                self.regions = ['us-east-1']
                print("✓ Using default: us-east-1")
            else:
                # Check if input is numbers or region names
                if all(part.strip().isdigit() for part in region_input.split(',')):
                    # Numbers provided
                    indices = [int(x.strip()) - 1 for x in region_input.split(',')]
                    self.regions = [common_regions[i] for i in indices if 0 <= i < len(common_regions)]
                else:
                    # Region names provided
                    self.regions = [r.strip() for r in region_input.split(',')]
                
                print(f"✓ Will analyze {len(self.regions)} region(s): {', '.join(self.regions)}")
        
        print()
        
        # Output Options
        print("="*80)
        print("OUTPUT OPTIONS")
        print("="*80)
        print("\nOutput formats will be generated:")
        print("  ✓ Console output (always generated)")
        print("  ✓ Text file report")
        if EXCEL_AVAILABLE:
            print("  ✓ Excel workbook (single file with multiple tabs)")
        else:
            print("  ✗ Excel not available (install openpyxl for Excel support)")
        
        self.output_prefix = input("\nEnter output filename prefix (default: storage_analysis): ").strip() or "storage_analysis"
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.output_prefix = f"{self.output_prefix}_{timestamp}"
        
        print(f"✓ Output files will be saved with prefix: {self.output_prefix}")
        print()
        
        # Confirmation
        print("="*80)
        print("ANALYSIS CONFIGURATION SUMMARY")
        print("="*80)
        print(f"AWS Profile:     {self.profile_name}")
        print(f"AWS Account:     {self.account_id}" + (f" ({self.account_alias})" if self.account_alias else ""))
        print(f"Regions:         {', '.join(self.regions)}")
        print(f"Output Prefix:   {self.output_prefix}")
        print("="*80)
        print()
        
        confirm = input("Start analysis? (yes/no) [default: yes]: ").strip().lower() or 'yes'
        
        if confirm not in ['y', 'yes']:
            print("\nAnalysis cancelled.")
            sys.exit(0)
        
        print("\n" + "="*80)
        print("STARTING ANALYSIS...")
        print("="*80)
        print()
    
    def _get_all_regions(self):
        """Get all available AWS regions"""
        try:
            ec2 = self.session.client('ec2', region_name='us-east-1')
            regions = ec2.describe_regions()['Regions']
            return [region['RegionName'] for region in regions]
        except Exception as e:
            print(f"Error fetching regions: {str(e)}")
            return ['us-east-1']
    
    def analyze_ec2_storage(self, region):
        """Analyze EC2 EBS volume storage in a region"""
        print(f"  → Analyzing EC2 volumes in {region}...")
        ec2 = self.session.client('ec2', region_name=region)
        total_gb = 0
        volume_count = 0
        volumes_detail = []
        
        try:
            paginator = ec2.get_paginator('describe_volumes')
            for page in paginator.paginate():
                for volume in page['Volumes']:
                    size_gb = volume['Size']
                    total_gb += size_gb
                    volume_count += 1
                    
                    # Get volume name from tags
                    volume_name = ''
                    if 'Tags' in volume:
                        for tag in volume['Tags']:
                            if tag['Key'] == 'Name':
                                volume_name = tag['Value']
                                break
                    
                    volumes_detail.append({
                        'volume_id': volume['VolumeId'],
                        'name': volume_name,
                        'size_gb': size_gb,
                        'state': volume['State'],
                        'volume_type': volume['VolumeType'],
                        'availability_zone': volume['AvailabilityZone'],
                        'attached_to': volume['Attachments'][0]['InstanceId'] if volume['Attachments'] else 'Not attached'
                    })
            
            print(f"    ✓ Found {volume_count} volumes totaling {total_gb} GB")
            return total_gb, volume_count, volumes_detail
            
        except ClientError as e:
            print(f"    ✗ Error analyzing EC2 volumes: {e}")
            return 0, 0, []
    
    def analyze_rds_storage(self, region):
        """Analyze RDS storage in a region"""
        print(f"  → Analyzing RDS instances in {region}...")
        rds = self.session.client('rds', region_name=region)
        total_gb = 0
        instance_count = 0
        instances_detail = []
        
        try:
            paginator = rds.get_paginator('describe_db_instances')
            for page in paginator.paginate():
                for db in page['DBInstances']:
                    size_gb = db['AllocatedStorage']
                    total_gb += size_gb
                    instance_count += 1
                    
                    instances_detail.append({
                        'db_identifier': db['DBInstanceIdentifier'],
                        'db_arn': db['DBInstanceArn'],
                        'engine': db['Engine'],
                        'engine_version': db['EngineVersion'],
                        'size_gb': size_gb,
                        'storage_type': db.get('StorageType', 'N/A'),
                        'status': db['DBInstanceStatus'],
                        'multi_az': db.get('MultiAZ', False)
                    })
            
            print(f"    ✓ Found {instance_count} RDS instances totaling {total_gb} GB")
            return total_gb, instance_count, instances_detail
            
        except ClientError as e:
            print(f"    ✗ Error analyzing RDS instances: {e}")
            return 0, 0, []
    
    def analyze_dynamodb_storage(self, region):
        """Analyze DynamoDB storage in a region"""
        print(f"  → Analyzing DynamoDB tables in {region}...")
        dynamodb = self.session.client('dynamodb', region_name=region)
        total_bytes = 0
        table_count = 0
        tables_detail = []
        
        try:
            paginator = dynamodb.get_paginator('list_tables')
            table_names = []
            for page in paginator.paginate():
                table_names.extend(page['TableNames'])
            
            for table_name in table_names:
                try:
                    table_info = dynamodb.describe_table(TableName=table_name)
                    table = table_info['Table']
                    size_bytes = table.get('TableSizeBytes', 0)
                    total_bytes += size_bytes
                    table_count += 1
                    
                    tables_detail.append({
                        'table_name': table_name,
                        'table_arn': table['TableArn'],
                        'size_gb': round(size_bytes / (1024**3), 4),
                        'item_count': table.get('ItemCount', 0),
                        'status': table['TableStatus'],
                        'billing_mode': table.get('BillingModeSummary', {}).get('BillingMode', 'PROVISIONED')
                    })
                except Exception as e:
                    print(f"    ⚠ Warning: Could not describe table {table_name}: {e}")
            
            total_gb = round(total_bytes / (1024**3), 2)
            print(f"    ✓ Found {table_count} DynamoDB tables totaling {total_gb} GB")
            return total_gb, table_count, tables_detail
            
        except ClientError as e:
            print(f"    ✗ Error analyzing DynamoDB tables: {e}")
            return 0, 0, []
    
    def analyze_aws_backups(self, region):
        """Analyze AWS Backup vaults and backup storage in a region"""
        print(f"  → Analyzing AWS Backups in {region}...")
        backup_client = self.session.client('backup', region_name=region)
        ec2 = self.session.client('ec2', region_name=region)
        rds = self.session.client('rds', region_name=region)
        dynamodb = self.session.client('dynamodb', region_name=region)
        
        total_backup_size_gb = 0
        total_source_size_gb = 0
        backup_count = 0
        backups_detail = []
        vault_summary = defaultdict(lambda: {'backup_count': 0, 'backup_size_gb': 0, 'source_size_gb': 0})
        
        try:
            # Get all backup vaults
            vaults_response = backup_client.list_backup_vaults()
            vaults = vaults_response.get('BackupVaultList', [])
            
            if not vaults:
                print(f"    ℹ No backup vaults found in {region}")
                return 0, 0, 0, backups_detail, dict(vault_summary)
            
            for vault in vaults:
                vault_name = vault['BackupVaultName']
                
                try:
                    # List recovery points (backups) in this vault
                    paginator = backup_client.get_paginator('list_recovery_points_by_backup_vault')
                    
                    for page in paginator.paginate(BackupVaultName=vault_name):
                        for recovery_point in page.get('RecoveryPoints', []):
                            backup_size_bytes = recovery_point.get('BackupSizeInBytes', 0)
                            backup_size_gb = round(backup_size_bytes / (1024**3), 4)
                            
                            resource_arn = recovery_point.get('ResourceArn', '')
                            resource_type = recovery_point.get('ResourceType', 'Unknown')
                            
                            # Get source resource size
                            source_size_gb = self._get_source_resource_size(
                                resource_arn, 
                                resource_type, 
                                region, 
                                ec2, 
                                rds, 
                                dynamodb
                            )
                            
                            total_backup_size_gb += backup_size_gb
                            total_source_size_gb += source_size_gb
                            backup_count += 1
                            
                            # Update vault summary
                            vault_summary[vault_name]['backup_count'] += 1
                            vault_summary[vault_name]['backup_size_gb'] += backup_size_gb
                            vault_summary[vault_name]['source_size_gb'] += source_size_gb
                            
                            backups_detail.append({
                                'vault_name': vault_name,
                                'recovery_point_arn': recovery_point.get('RecoveryPointArn', ''),
                                'resource_arn': resource_arn,
                                'resource_type': resource_type,
                                'resource_id': self._extract_resource_id(resource_arn, resource_type),
                                'source_size_gb': source_size_gb,
                                'backup_size_gb': backup_size_gb,
                                'creation_date': recovery_point.get('CreationDate', '').isoformat() if recovery_point.get('CreationDate') else 'N/A',
                                'status': recovery_point.get('Status', 'Unknown'),
                                'lifecycle': recovery_point.get('Lifecycle', {}).get('DeleteAfterDays', 'Never')
                            })
                            
                except ClientError as e:
                    if e.response['Error']['Code'] == 'AccessDeniedException':
                        print(f"    ⚠ Warning: Access denied to vault {vault_name}")
                    else:
                        print(f"    ⚠ Warning: Error accessing vault {vault_name}: {e}")
                    continue
            
            print(f"    ✓ Found {backup_count} backups in {len(vaults)} vaults")
            print(f"      Backup Storage: {total_backup_size_gb:,.2f} GB")
            print(f"      Source Storage: {total_source_size_gb:,.2f} GB")
            
            return total_backup_size_gb, total_source_size_gb, backup_count, backups_detail, dict(vault_summary)
            
        except ClientError as e:
            if e.response['Error']['Code'] == 'AccessDeniedException':
                print(f"    ⚠ Warning: Access denied to AWS Backup service in {region}")
            else:
                print(f"    ✗ Error analyzing AWS Backups: {e}")
            return 0, 0, 0, [], {}
    
    def _extract_resource_id(self, resource_arn, resource_type):
        """Extract resource ID from ARN"""
        if not resource_arn:
            return 'N/A'
        
        try:
            if resource_type == 'EBS':
                # arn:aws:ec2:region:account:volume/vol-xxxxx
                return resource_arn.split('/')[-1]
            elif resource_type == 'RDS':
                # arn:aws:rds:region:account:db:instance-name
                return resource_arn.split(':')[-1]
            elif resource_type == 'DynamoDB':
                # arn:aws:dynamodb:region:account:table/table-name
                return resource_arn.split('/')[-1]
            elif resource_type == 'EC2':
                # arn:aws:ec2:region:account:instance/i-xxxxx
                return resource_arn.split('/')[-1]
            else:
                # Try to get the last part
                if '/' in resource_arn:
                    return resource_arn.split('/')[-1]
                elif ':' in resource_arn:
                    return resource_arn.split(':')[-1]
                return resource_arn
        except:
            return resource_arn
    
    def _get_source_resource_size(self, resource_arn, resource_type, region, ec2_client, rds_client, dynamodb_client):
        """Get the size of the source resource being backed up"""
        try:
            if resource_type == 'EBS':
                volume_id = resource_arn.split('/')[-1]
                response = ec2_client.describe_volumes(VolumeIds=[volume_id])
                if response['Volumes']:
                    return response['Volumes'][0]['Size']
            
            elif resource_type == 'EC2':
                instance_id = resource_arn.split('/')[-1]
                response = ec2_client.describe_instances(InstanceIds=[instance_id])
                if response['Reservations']:
                    total_size = 0
                    for reservation in response['Reservations']:
                        for instance in reservation['Instances']:
                            for bdm in instance.get('BlockDeviceMappings', []):
                                if 'Ebs' in bdm:
                                    volume_id = bdm['Ebs']['VolumeId']
                                    vol_response = ec2_client.describe_volumes(VolumeIds=[volume_id])
                                    if vol_response['Volumes']:
                                        total_size += vol_response['Volumes'][0]['Size']
                    return total_size
            
            elif resource_type == 'RDS':
                db_identifier = resource_arn.split(':')[-1]
                response = rds_client.describe_db_instances(DBInstanceIdentifier=db_identifier)
                if response['DBInstances']:
                    return response['DBInstances'][0]['AllocatedStorage']
            
            elif resource_type == 'DynamoDB':
                table_name = resource_arn.split('/')[-1]
                response = dynamodb_client.describe_table(TableName=table_name)
                if response['Table']:
                    size_bytes = response['Table'].get('TableSizeBytes', 0)
                    return round(size_bytes / (1024**3), 4)
            
        except Exception as e:
            # Resource might have been deleted or we don't have permission
            pass
        
        return 0
    
    def analyze_region(self, region):
        """Analyze storage in a specific region"""
        print(f"\n{'='*80}")
        print(f"Analyzing Region: {region}")
        print(f"{'='*80}")
        
        ec2_total, ec2_count, ec2_details = self.analyze_ec2_storage(region)
        rds_total, rds_count, rds_details = self.analyze_rds_storage(region)
        dynamo_total, dynamo_count, dynamo_details = self.analyze_dynamodb_storage(region)
        backup_size, source_size, backup_count, backup_details, vault_summary = self.analyze_aws_backups(region)
        
        total_storage = ec2_total + rds_total + dynamo_total
        
        result = {
            'region': region,
            'ec2_storage_gb': ec2_total,
            'ec2_volume_count': ec2_count,
            'ec2_details': ec2_details,
            'rds_storage_gb': rds_total,
            'rds_instance_count': rds_count,
            'rds_details': rds_details,
            'dynamodb_storage_gb': dynamo_total,
            'dynamodb_table_count': dynamo_count,
            'dynamodb_details': dynamo_details,
            'backup_storage_gb': backup_size,
            'backup_source_storage_gb': source_size,
            'backup_count': backup_count,
            'backup_details': backup_details,
            'vault_summary': vault_summary,
            'total_storage_gb': total_storage
        }
        
        self.results.append(result)
        
        print(f"\n  Summary for {region}:")
        print(f"    EC2 Storage:      {ec2_total:,.2f} GB ({ec2_count} volumes)")
        print(f"    RDS Storage:      {rds_total:,.2f} GB ({rds_count} instances)")
        print(f"    DynamoDB Storage: {dynamo_total:,.2f} GB ({dynamo_count} tables)")
        print(f"    AWS Backups:      {backup_size:,.2f} GB ({backup_count} recovery points)")
        print(f"    Backup Source:    {source_size:,.2f} GB (original resource sizes)")
        print(f"    Total Storage:    {total_storage:,.2f} GB")
        
        return result
    
    def run_analysis(self):
        """Run the storage analysis across all configured regions"""
        for region in self.regions:
            try:
                self.analyze_region(region)
            except Exception as e:
                print(f"\n✗ Error analyzing region {region}: {e}")
                continue
        
        self.print_summary()
        self.export_results()
    
    def print_summary(self):
        """Print overall summary of storage analysis"""
        print("\n" + "="*80)
        print("OVERALL STORAGE SUMMARY")
        print("="*80)
        
        total_ec2 = sum(r['ec2_storage_gb'] for r in self.results)
        total_rds = sum(r['rds_storage_gb'] for r in self.results)
        total_dynamodb = sum(r['dynamodb_storage_gb'] for r in self.results)
        total_backup_storage = sum(r['backup_storage_gb'] for r in self.results)
        total_backup_source = sum(r['backup_source_storage_gb'] for r in self.results)
        grand_total = sum(r['total_storage_gb'] for r in self.results)
        
        total_ec2_volumes = sum(r['ec2_volume_count'] for r in self.results)
        total_rds_instances = sum(r['rds_instance_count'] for r in self.results)
        total_dynamo_tables = sum(r['dynamodb_table_count'] for r in self.results)
        total_backups = sum(r['backup_count'] for r in self.results)
        
        print(f"\nAccount: {self.account_id}" + (f" ({self.account_alias})" if self.account_alias else ""))
        print(f"Regions Analyzed: {len(self.regions)}")
        print(f"\n{'Service':<25} {'Storage (GB)':<20} {'Resource Count':<20}")
        print("-" * 65)
        print(f"{'EC2 (EBS)':<25} {total_ec2:>15,.2f}     {total_ec2_volumes:>15,}")
        print(f"{'RDS':<25} {total_rds:>15,.2f}     {total_rds_instances:>15,}")
        print(f"{'DynamoDB':<25} {total_dynamodb:>15,.2f}     {total_dynamo_tables:>15,}")
        print(f"{'AWS Backup (stored)':<25} {total_backup_storage:>15,.2f}     {total_backups:>15,}")
        print(f"{'AWS Backup (source)':<25} {total_backup_source:>15,.2f}")
        print("-" * 65)
        print(f"{'TOTAL (Active Storage)':<25} {grand_total:>15,.2f}")
        print(f"{'TOTAL (inc. Backups)':<25} {grand_total + total_backup_storage:>15,.2f}")
        print()
        
        # Backup efficiency
        if total_backup_source > 0:
            compression_ratio = (total_backup_storage / total_backup_source) * 100
            print(f"Backup Efficiency: {compression_ratio:.1f}% (backup size vs source size)")
            print()
        
        # Region breakdown
        if len(self.regions) > 1:
            print("\nStorage by Region:")
            print(f"{'Region':<20} {'Active (GB)':<20} {'Backup (GB)':<20} {'Total (GB)':<20}")
            print("-" * 80)
            for result in sorted(self.results, key=lambda x: x['total_storage_gb'], reverse=True):
                active_storage = result['total_storage_gb']
                backup_storage = result['backup_storage_gb']
                total = active_storage + backup_storage
                print(f"{result['region']:<20} {active_storage:>15,.2f}     {backup_storage:>15,.2f}     {total:>15,.2f}")
            print()
    
    def export_to_text(self):
        """Export results to text file"""
        filename = f"{self.output_prefix}.txt"
        
        try:
            with open(filename, 'w') as f:
                f.write("="*80 + "\n")
                f.write("AWS STORAGE ANALYSIS REPORT\n")
                f.write("="*80 + "\n\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Account: {self.account_id}")
                if self.account_alias:
                    f.write(f" ({self.account_alias})")
                f.write(f"\nProfile: {self.profile_name}\n")
                f.write(f"Regions: {', '.join(self.regions)}\n\n")
                
                # Overall Summary
                f.write("="*80 + "\n")
                f.write("OVERALL SUMMARY\n")
                f.write("="*80 + "\n\n")
                
                total_ec2 = sum(r['ec2_storage_gb'] for r in self.results)
                total_rds = sum(r['rds_storage_gb'] for r in self.results)
                total_dynamodb = sum(r['dynamodb_storage_gb'] for r in self.results)
                total_backup_storage = sum(r['backup_storage_gb'] for r in self.results)
                total_backup_source = sum(r['backup_source_storage_gb'] for r in self.results)
                grand_total = sum(r['total_storage_gb'] for r in self.results)
                
                f.write(f"EC2 Storage Total (GB):              {total_ec2:,.2f}\n")
                f.write(f"RDS Storage Total (GB):              {total_rds:,.2f}\n")
                f.write(f"DynamoDB Storage Total (GB):         {total_dynamodb:,.2f}\n")
                f.write(f"AWS Backup Storage Total (GB):       {total_backup_storage:,.2f}\n")
                f.write(f"AWS Backup Source Storage (GB):      {total_backup_source:,.2f}\n")
                f.write(f"{'─'*40}\n")
                f.write(f"TOTAL ACTIVE STORAGE (GB):           {grand_total:,.2f}\n")
                f.write(f"TOTAL WITH BACKUPS (GB):             {grand_total + total_backup_storage:,.2f}\n\n")
                
                # Regional Details
                for result in self.results:
                    f.write("\n" + "="*80 + "\n")
                    f.write(f"Region: {result['region']}\n")
                    f.write("="*80 + "\n\n")
                    
                    f.write(f"EC2 Storage:      {result['ec2_storage_gb']:,.2f} GB ({result['ec2_volume_count']} volumes)\n")
                    f.write(f"RDS Storage:      {result['rds_storage_gb']:,.2f} GB ({result['rds_instance_count']} instances)\n")
                    f.write(f"DynamoDB Storage: {result['dynamodb_storage_gb']:,.2f} GB ({result['dynamodb_table_count']} tables)\n")
                    f.write(f"AWS Backups:      {result['backup_storage_gb']:,.2f} GB ({result['backup_count']} recovery points)\n")
                    f.write(f"Backup Source:    {result['backup_source_storage_gb']:,.2f} GB\n")
                    f.write(f"Total:            {result['total_storage_gb']:,.2f} GB\n\n")
                    
                    # Backup Vault Summary
                    if result['vault_summary']:
                        f.write("  Backup Vaults:\n")
                        f.write("  " + "-"*76 + "\n")
                        for vault_name, vault_data in result['vault_summary'].items():
                            f.write(f"    Vault:        {vault_name}\n")
                            f.write(f"    Backups:      {vault_data['backup_count']}\n")
                            f.write(f"    Backup Size:  {vault_data['backup_size_gb']:,.2f} GB\n")
                            f.write(f"    Source Size:  {vault_data['source_size_gb']:,.2f} GB\n")
                            f.write("\n")
                    
                    # EC2 Details
                    if result['ec2_details']:
                        f.write("  EC2 Volumes:\n")
                        f.write("  " + "-"*76 + "\n")
                        for vol in result['ec2_details']:
                            f.write(f"    Volume ID: {vol['volume_id']}\n")
                            if vol['name']:
                                f.write(f"    Name:      {vol['name']}\n")
                            f.write(f"    Size:      {vol['size_gb']} GB\n")
                            f.write(f"    Type:      {vol['volume_type']}\n")
                            f.write(f"    State:     {vol['state']}\n")
                            f.write(f"    AZ:        {vol['availability_zone']}\n")
                            f.write(f"    Attached:  {vol['attached_to']}\n")
                            f.write("\n")
                    
                    # RDS Details
                    if result['rds_details']:
                        f.write("  RDS Instances:\n")
                        f.write("  " + "-"*76 + "\n")
                        for db in result['rds_details']:
                            f.write(f"    Identifier:  {db['db_identifier']}\n")
                            f.write(f"    Engine:      {db['engine']} {db['engine_version']}\n")
                            f.write(f"    Size:        {db['size_gb']} GB\n")
                            f.write(f"    Type:        {db['storage_type']}\n")
                            f.write(f"    Status:      {db['status']}\n")
                            f.write(f"    Multi-AZ:    {db['multi_az']}\n")
                            f.write("\n")
                    
                    # DynamoDB Details
                    if result['dynamodb_details']:
                        f.write("  DynamoDB Tables:\n")
                        f.write("  " + "-"*76 + "\n")
                        for table in result['dynamodb_details']:
                            f.write(f"    Table:        {table['table_name']}\n")
                            f.write(f"    Size:         {table['size_gb']} GB\n")
                            f.write(f"    Items:        {table['item_count']:,}\n")
                            f.write(f"    Status:       {table['status']}\n")
                            f.write(f"    Billing Mode: {table['billing_mode']}\n")
                            f.write("\n")
                    
                    # Backup Details
                    if result['backup_details']:
                        f.write("  AWS Backup Recovery Points:\n")
                        f.write("  " + "-"*76 + "\n")
                        for backup in result['backup_details']:
                            f.write(f"    Vault:         {backup['vault_name']}\n")
                            f.write(f"    Resource Type: {backup['resource_type']}\n")
                            f.write(f"    Resource ID:   {backup['resource_id']}\n")
                            f.write(f"    Source Size:   {backup['source_size_gb']:.2f} GB\n")
                            f.write(f"    Backup Size:   {backup['backup_size_gb']:.2f} GB\n")
                            f.write(f"    Created:       {backup['creation_date']}\n")
                            f.write(f"    Status:        {backup['status']}\n")
                            f.write(f"    Retention:     {backup['lifecycle']}\n")
                            f.write("\n")
            
            print(f"✓ Text report exported to: {filename}")
            
        except Exception as e:
            print(f"✗ Error exporting to text file: {e}")
    
    def export_to_excel(self):
        """Export results to Excel workbook with multiple sheets"""
        if not EXCEL_AVAILABLE:
            print("✗ Excel export skipped (openpyxl not available)")
            return
        
        filename = f"{self.output_prefix}.xlsx"
        
        try:
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            section_font = Font(bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Summary Sheet - Executive Report Style
            ws_summary = wb.create_sheet("Executive Summary")
            
            # Report Header
            ws_summary['A1'] = "AWS STORAGE ANALYSIS"
            ws_summary['A1'].font = Font(size=16, bold=True, color="366092")
            ws_summary.merge_cells('A1:F1')
            
            ws_summary['A2'] = "Executive Summary Report"
            ws_summary['A2'].font = Font(size=12, italic=True)
            ws_summary.merge_cells('A2:F2')
            
            row = 4
            ws_summary[f'A{row}'] = f"Report Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            ws_summary[f'A{row}'].font = Font(size=10)
            
            row += 1
            ws_summary[f'A{row}'] = f"AWS Account: {self.account_id}" + (f" ({self.account_alias})" if self.account_alias else "")
            ws_summary[f'A{row}'].font = Font(size=10)
            
            row += 1
            ws_summary[f'A{row}'] = f"Regions Analyzed: {', '.join(self.regions)}"
            ws_summary[f'A{row}'].font = Font(size=10)
            
            row += 1
            ws_summary[f'A{row}'] = f"Analysis Profile: {self.profile_name}"
            ws_summary[f'A{row}'].font = Font(size=10)
            
            # Executive Summary Section
            row += 3
            ws_summary[f'A{row}'] = "EXECUTIVE SUMMARY"
            ws_summary[f'A{row}'].font = section_font
            ws_summary[f'A{row}'].fill = section_fill
            ws_summary.merge_cells(f'A{row}:F{row}')
            
            row += 2
            ws_summary[f'A{row}'] = "This report provides a comprehensive analysis of storage utilization across your AWS infrastructure."
            ws_summary.merge_cells(f'A{row}:F{row}')
            ws_summary[f'A{row}'].alignment = Alignment(wrap_text=True)
            
            row += 1
            ws_summary[f'A{row}'] = "The analysis covers active storage resources (EC2 EBS volumes, RDS databases, DynamoDB tables) and backup storage (AWS Backup service)."
            ws_summary.merge_cells(f'A{row}:F{row}')
            ws_summary[f'A{row}'].alignment = Alignment(wrap_text=True)
            
            # Storage Overview Section
            row += 3
            ws_summary[f'A{row}'] = "STORAGE OVERVIEW"
            ws_summary[f'A{row}'].font = section_font
            ws_summary[f'A{row}'].fill = section_fill
            ws_summary.merge_cells(f'A{row}:C{row}')
            
            row += 2
            ws_summary[f'A{row}'] = "Service"
            ws_summary[f'B{row}'] = "Total Storage (GB)"
            ws_summary[f'C{row}'] = "Resource Count"
            
            for col in ['A', 'B', 'C']:
                ws_summary[f'{col}{row}'].fill = header_fill
                ws_summary[f'{col}{row}'].font = header_font
                ws_summary[f'{col}{row}'].border = border
                ws_summary[f'{col}{row}'].alignment = Alignment(horizontal='center')
            
            row += 1
            total_ec2 = sum(r['ec2_storage_gb'] for r in self.results)
            total_rds = sum(r['rds_storage_gb'] for r in self.results)
            total_dynamodb = sum(r['dynamodb_storage_gb'] for r in self.results)
            total_backup_storage = sum(r['backup_storage_gb'] for r in self.results)
            total_backup_source = sum(r['backup_source_storage_gb'] for r in self.results)
            
            ws_summary[f'A{row}'] = "EC2 (EBS Volumes)"
            ws_summary[f'B{row}'] = round(total_ec2, 2)
            ws_summary[f'C{row}'] = sum(r['ec2_volume_count'] for r in self.results)
            ws_summary[f'B{row}'].number_format = '#,##0.00'
            
            row += 1
            ws_summary[f'A{row}'] = "RDS (Databases)"
            ws_summary[f'B{row}'] = round(total_rds, 2)
            ws_summary[f'C{row}'] = sum(r['rds_instance_count'] for r in self.results)
            ws_summary[f'B{row}'].number_format = '#,##0.00'
            
            row += 1
            ws_summary[f'A{row}'] = "DynamoDB (Tables)"
            ws_summary[f'B{row}'] = round(total_dynamodb, 2)
            ws_summary[f'C{row}'] = sum(r['dynamodb_table_count'] for r in self.results)
            ws_summary[f'B{row}'].number_format = '#,##0.00'
            
            row += 1
            ws_summary[f'A{row}'] = "AWS Backup (Stored)"
            ws_summary[f'B{row}'] = round(total_backup_storage, 2)
            ws_summary[f'C{row}'] = sum(r['backup_count'] for r in self.results)
            ws_summary[f'B{row}'].number_format = '#,##0.00'
            
            row += 1
            ws_summary[f'A{row}'] = "AWS Backup (Source)"
            ws_summary[f'B{row}'] = round(total_backup_source, 2)
            ws_summary[f'C{row}'] = "N/A"
            ws_summary[f'B{row}'].number_format = '#,##0.00'
            
            row += 1
            for col in ['A', 'B', 'C']:
                ws_summary[f'{col}{row}'].border = Border(top=Side(style='double'))
            
            ws_summary[f'A{row}'] = "TOTAL (Active Storage)"
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'] = round(total_ec2 + total_rds + total_dynamodb, 2)
            ws_summary[f'B{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'].number_format = '#,##0.00'
            ws_summary[f'C{row}'].font = Font(bold=True)
            
            row += 1
            ws_summary[f'A{row}'] = "TOTAL (inc. Backups)"
            ws_summary[f'A{row}'].font = Font(bold=True, color="C00000")
            ws_summary[f'B{row}'] = round(total_ec2 + total_rds + total_dynamodb + total_backup_storage, 2)
            ws_summary[f'B{row}'].font = Font(bold=True, color="C00000")
            ws_summary[f'B{row}'].number_format = '#,##0.00'
            ws_summary[f'C{row}'].font = Font(bold=True)
            
            # Key Metrics Section
            row += 3
            ws_summary[f'A{row}'] = "KEY METRICS"
            ws_summary[f'A{row}'].font = section_font
            ws_summary[f'A{row}'].fill = section_fill
            ws_summary.merge_cells(f'A{row}:C{row}')
            
            row += 2
            if total_backup_source > 0:
                compression_ratio = (total_backup_storage / total_backup_source) * 100
                ws_summary[f'A{row}'] = "Backup Efficiency:"
                ws_summary[f'A{row}'].font = Font(bold=True)
                ws_summary[f'B{row}'] = f"{compression_ratio:.1f}%"
                ws_summary[f'C{row}'] = "(Backup size vs source size - lower is better)"
                ws_summary.merge_cells(f'C{row}:F{row}')
                
                row += 1
            
            ws_summary[f'A{row}'] = "Total Regions:"
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'] = len(self.regions)
            
            row += 1
            ws_summary[f'A{row}'] = "Total Resources:"
            ws_summary[f'A{row}'].font = Font(bold=True)
            total_resources = (sum(r['ec2_volume_count'] for r in self.results) + 
                             sum(r['rds_instance_count'] for r in self.results) + 
                             sum(r['dynamodb_table_count'] for r in self.results))
            ws_summary[f'B{row}'] = total_resources
            
            row += 1
            ws_summary[f'A{row}'] = "Total Backup Points:"
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'] = sum(r['backup_count'] for r in self.results)
            
            # Regional breakdown
            row += 3
            ws_summary[f'A{row}'] = "STORAGE BY REGION"
            ws_summary[f'A{row}'].font = section_font
            ws_summary[f'A{row}'].fill = section_fill
            ws_summary.merge_cells(f'A{row}:F{row}')
            
            row += 2
            ws_summary[f'A{row}'] = "Region"
            ws_summary[f'B{row}'] = "EC2 (GB)"
            ws_summary[f'C{row}'] = "RDS (GB)"
            ws_summary[f'D{row}'] = "DynamoDB (GB)"
            ws_summary[f'E{row}'] = "Backup (GB)"
            ws_summary[f'F{row}'] = "Total (GB)"
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws_summary[f'{col}{row}'].fill = header_fill
                ws_summary[f'{col}{row}'].font = header_font
                ws_summary[f'{col}{row}'].border = border
                ws_summary[f'{col}{row}'].alignment = Alignment(horizontal='center')
            
            for result in sorted(self.results, key=lambda x: x['total_storage_gb'] + x['backup_storage_gb'], reverse=True):
                row += 1
                ws_summary[f'A{row}'] = result['region']
                ws_summary[f'B{row}'] = round(result['ec2_storage_gb'], 2)
                ws_summary[f'C{row}'] = round(result['rds_storage_gb'], 2)
                ws_summary[f'D{row}'] = round(result['dynamodb_storage_gb'], 2)
                ws_summary[f'E{row}'] = round(result['backup_storage_gb'], 2)
                ws_summary[f'F{row}'] = round(result['total_storage_gb'] + result['backup_storage_gb'], 2)
                
                for col in ['B', 'C', 'D', 'E', 'F']:
                    ws_summary[f'{col}{row}'].number_format = '#,##0.00'
            
            # Report Tabs Description Section
            row += 4
            ws_summary[f'A{row}'] = "DETAILED ANALYSIS TABS"
            ws_summary[f'A{row}'].font = section_font
            ws_summary[f'A{row}'].fill = section_fill
            ws_summary.merge_cells(f'A{row}:F{row}')
            
            row += 2
            ws_summary[f'A{row}'] = "Tab Name"
            ws_summary[f'A{row}'].font = Font(bold=True)
            ws_summary[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws_summary[f'B{row}'] = "Description"
            ws_summary[f'B{row}'].font = Font(bold=True)
            ws_summary[f'B{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            ws_summary.merge_cells(f'B{row}:F{row}')
            
            # Tab descriptions
            tab_descriptions = [
                {
                    'name': 'EC2 Volumes',
                    'description': 'Detailed inventory of all EBS volumes across regions. Includes volume IDs, names (tags), sizes, types (gp3, io1, etc.), states (in-use, available), availability zones, and attachment information. Use this to identify unused volumes, oversized volumes, or opportunities for volume type optimization.'
                },
                {
                    'name': 'RDS Instances',
                    'description': 'Complete listing of RDS database instances with storage allocations. Shows database identifiers, engine types (MySQL, PostgreSQL, etc.), versions, allocated storage, storage types (gp2, io1), operational status, and Multi-AZ configuration. Critical for database capacity planning and cost optimization.'
                },
                {
                    'name': 'DynamoDB Tables',
                    'description': 'Comprehensive view of DynamoDB tables and their storage consumption. Displays table names, actual storage sizes, item counts, table status, and billing modes (On-Demand vs Provisioned). Essential for understanding NoSQL database footprint and usage patterns.'
                },
                {
                    'name': 'AWS Backups',
                    'description': 'Detailed recovery point inventory showing all AWS Backup snapshots and their storage impact. Lists backup vaults, resource types being protected (EC2, RDS, EBS, DynamoDB), original resource sizes vs compressed backup sizes, creation dates, status, and retention policies. Use this to understand backup costs and validate protection coverage.'
                },
                {
                    'name': 'Backup Vaults',
                    'description': 'Aggregated view of backup vaults with summary statistics. Shows total backup counts, cumulative backup storage, and total source storage per vault. Helps identify backup vault utilization and storage efficiency by comparing backup size to original resource size.'
                }
            ]
            
            for tab_info in tab_descriptions:
                row += 1
                ws_summary[f'A{row}'] = tab_info['name']
                ws_summary[f'A{row}'].font = Font(bold=True, color="366092")
                ws_summary[f'A{row}'].alignment = Alignment(vertical='top')
                
                ws_summary[f'B{row}'] = tab_info['description']
                ws_summary[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                ws_summary.merge_cells(f'B{row}:F{row}')
                ws_summary.row_dimensions[row].height = 45
            
            # Key Insights Section
            row += 3
            ws_summary[f'A{row}'] = "KEY INSIGHTS & RECOMMENDATIONS"
            ws_summary[f'A{row}'].font = section_font
            ws_summary[f'A{row}'].fill = section_fill
            ws_summary.merge_cells(f'A{row}:F{row}')
            
            row += 2
            
            # Generate insights based on data
            insights = []
            
            # Check for unattached volumes
            unattached_volumes = 0
            unattached_storage = 0
            for result in self.results:
                for vol in result['ec2_details']:
                    if vol['attached_to'] == 'Not attached':
                        unattached_volumes += 1
                        unattached_storage += vol['size_gb']
            
            if unattached_volumes > 0:
                insights.append(f"• Found {unattached_volumes} unattached EBS volumes totaling {unattached_storage:,.2f} GB. Consider deleting unused volumes to reduce costs.")
            
            # Check backup efficiency
            if total_backup_source > 0:
                compression_ratio = (total_backup_storage / total_backup_source) * 100
                if compression_ratio > 80:
                    insights.append(f"• Backup compression ratio is {compression_ratio:.1f}%. This is relatively high - consider reviewing backup policies or using alternative backup strategies for large datasets.")
                elif compression_ratio < 30:
                    insights.append(f"• Backup compression ratio is {compression_ratio:.1f}%. Excellent backup efficiency achieved through deduplication and compression.")
            
            # Check for backup coverage
            total_active_resources = (sum(r['ec2_volume_count'] for r in self.results) + 
                                    sum(r['rds_instance_count'] for r in self.results) + 
                                    sum(r['dynamodb_table_count'] for r in self.results))
            
            if total_backup_source < (total_ec2 + total_rds + total_dynamodb) * 0.3:
                insights.append(f"• Only {(total_backup_source / (total_ec2 + total_rds + total_dynamodb) * 100):.1f}% of active storage is being backed up. Review backup policies to ensure adequate data protection.")
            
            # Check regional distribution
            if len(self.regions) > 1:
                largest_region = max(self.results, key=lambda x: x['total_storage_gb'])
                if largest_region['total_storage_gb'] > sum(r['total_storage_gb'] for r in self.results) * 0.7:
                    insights.append(f"• {largest_region['region']} contains {(largest_region['total_storage_gb'] / sum(r['total_storage_gb'] for r in self.results) * 100):.1f}% of total storage. Consider geographic distribution for disaster recovery.")
            
            # DynamoDB insights
            if total_dynamodb > 100:
                insights.append(f"• DynamoDB storage is {total_dynamodb:,.2f} GB. Review table usage patterns and consider enabling Point-in-Time Recovery for critical tables.")
            
            if not insights:
                insights.append("• Storage distribution appears well-balanced across services and regions.")
                insights.append("• Continue monitoring storage growth trends and backup coverage.")
            
            for insight in insights:
                row += 1
                ws_summary[f'A{row}'] = insight
                ws_summary.merge_cells(f'A{row}:F{row}')
                ws_summary[f'A{row}'].alignment = Alignment(wrap_text=True)
                ws_summary.row_dimensions[row].height = 30
            
            # Footer
            row += 3
            ws_summary[f'A{row}'] = "Notes:"
            ws_summary[f'A{row}'].font = Font(bold=True, size=9)
            row += 1
            ws_summary[f'A{row}'] = "• All storage sizes are reported in Gigabytes (GB)"
            ws_summary[f'A{row}'].font = Font(size=9, italic=True)
            ws_summary.merge_cells(f'A{row}:F{row}')
            row += 1
            ws_summary[f'A{row}'] = "• Backup (Source) refers to the original size of resources being backed up"
            ws_summary[f'A{row}'].font = Font(size=9, italic=True)
            ws_summary.merge_cells(f'A{row}:F{row}')
            row += 1
            ws_summary[f'A{row}'] = "• Backup (Stored) refers to the actual compressed/deduplicated backup storage consumed"
            ws_summary[f'A{row}'].font = Font(size=9, italic=True)
            ws_summary.merge_cells(f'A{row}:F{row}')
            row += 1
            ws_summary[f'A{row}'] = f"• Analysis performed using AWS profile: {self.profile_name}"
            ws_summary[f'A{row}'].font = Font(size=9, italic=True)
            ws_summary.merge_cells(f'A{row}:F{row}')
            
            # Auto-size columns
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 20
            ws_summary.column_dimensions['C'].width = 20
            ws_summary.column_dimensions['D'].width = 20
            ws_summary.column_dimensions['E'].width = 20
            ws_summary.column_dimensions['F'].width = 20
            
            # EC2 Details Sheet
            ws_ec2 = wb.create_sheet("EC2 Volumes")
            headers = ['Region', 'Volume ID', 'Name', 'Size (GB)', 'Type', 'State', 'AZ', 'Attached To']
            ws_ec2.append(headers)
            
            for col_num, _ in enumerate(headers, 1):
                cell = ws_ec2.cell(1, col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            for result in self.results:
                for vol in result['ec2_details']:
                    ws_ec2.append([
                        result['region'],
                        vol['volume_id'],
                        vol['name'],
                        vol['size_gb'],
                        vol['volume_type'],
                        vol['state'],
                        vol['availability_zone'],
                        vol['attached_to']
                    ])
            
            for col in ws_ec2.columns:
                ws_ec2.column_dimensions[col[0].column_letter].width = 18
            
            # RDS Details Sheet
            ws_rds = wb.create_sheet("RDS Instances")
            headers = ['Region', 'DB Identifier', 'Engine', 'Version', 'Size (GB)', 'Storage Type', 'Status', 'Multi-AZ']
            ws_rds.append(headers)
            
            for col_num, _ in enumerate(headers, 1):
                cell = ws_rds.cell(1, col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            for result in self.results:
                for db in result['rds_details']:
                    ws_rds.append([
                        result['region'],
                        db['db_identifier'],
                        db['engine'],
                        db['engine_version'],
                        db['size_gb'],
                        db['storage_type'],
                        db['status'],
                        db['multi_az']
                    ])
            
            for col in ws_rds.columns:
                ws_rds.column_dimensions[col[0].column_letter].width = 18
            
            # DynamoDB Details Sheet
            ws_dynamo = wb.create_sheet("DynamoDB Tables")
            headers = ['Region', 'Table Name', 'Size (GB)', 'Item Count', 'Status', 'Billing Mode']
            ws_dynamo.append(headers)
            
            for col_num, _ in enumerate(headers, 1):
                cell = ws_dynamo.cell(1, col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            for result in self.results:
                for table in result['dynamodb_details']:
                    ws_dynamo.append([
                        result['region'],
                        table['table_name'],
                        table['size_gb'],
                        table['item_count'],
                        table['status'],
                        table['billing_mode']
                    ])
            
            for col in ws_dynamo.columns:
                ws_dynamo.column_dimensions[col[0].column_letter].width = 20
            
            # AWS Backup Details Sheet
            ws_backup = wb.create_sheet("AWS Backups")
            headers = ['Region', 'Vault Name', 'Resource Type', 'Resource ID', 'Source Size (GB)', 
                      'Backup Size (GB)', 'Creation Date', 'Status', 'Retention']
            ws_backup.append(headers)
            
            for col_num, _ in enumerate(headers, 1):
                cell = ws_backup.cell(1, col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            for result in self.results:
                for backup in result['backup_details']:
                    ws_backup.append([
                        result['region'],
                        backup['vault_name'],
                        backup['resource_type'],
                        backup['resource_id'],
                        backup['source_size_gb'],
                        backup['backup_size_gb'],
                        backup['creation_date'],
                        backup['status'],
                        str(backup['lifecycle'])
                    ])
            
            for col in ws_backup.columns:
                ws_backup.column_dimensions[col[0].column_letter].width = 20
            
            # Backup Vault Summary Sheet
            ws_vault = wb.create_sheet("Backup Vaults")
            headers = ['Region', 'Vault Name', 'Backup Count', 'Backup Size (GB)', 'Source Size (GB)']
            ws_vault.append(headers)
            
            for col_num, _ in enumerate(headers, 1):
                cell = ws_vault.cell(1, col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            for result in self.results:
                for vault_name, vault_data in result['vault_summary'].items():
                    ws_vault.append([
                        result['region'],
                        vault_name,
                        vault_data['backup_count'],
                        vault_data['backup_size_gb'],
                        vault_data['source_size_gb']
                    ])
            
            for col in ws_vault.columns:
                ws_vault.column_dimensions[col[0].column_letter].width = 22
            
            wb.save(filename)
            print(f"✓ Excel report exported to: {filename}")
            
        except Exception as e:
            print(f"✗ Error exporting to Excel: {e}")
            import traceback
            traceback.print_exc()
    
    def export_results(self):
        """Export results to all configured formats"""
        print("\n" + "="*80)
        print("EXPORTING RESULTS")
        print("="*80)
        print()
        
        self.export_to_text()
        self.export_to_excel()
        
        print()


def main():
    """Main entry point"""
    analyzer = StorageAnalyzer()
    
    try:
        analyzer.setup_interactive()
        analyzer.run_analysis()
        
        print("="*80)
        print("ANALYSIS COMPLETE")
        print("="*80)
        print()
        
    except KeyboardInterrupt:
        print("\n\nAnalysis interrupted by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\n✗ Fatal error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
