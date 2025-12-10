#!/usr/bin/env python3
"""
AWS NAT Gateway Analyzer
Analyzes NAT Gateways in an AWS account and identifies resources using them.
"""

import boto3
import botocore.session
import csv
import sys
import time
from datetime import datetime
from collections import defaultdict
from botocore.exceptions import ClientError

# Try to import openpyxl for Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will not be available.")
    print("Install with: pip install openpyxl")

class NATGatewayAnalyzer:
    def __init__(self):
        """Initialize the analyzer"""
        self.session = None
        self.regions = []
        self.results = []
        self.account_id = None
        self.account_alias = None
        self.profile_name = None
        self.output_prefix = None
        
    def get_profile_and_account_info(self):
        """Prompt for AWS profile and confirm account details"""
        session = botocore.session.Session()
        profiles = session.available_profiles
        
        if not profiles:
            print("Error: No AWS profiles found. Please configure AWS CLI first.")
            sys.exit(1)
        
        print("\n" + "="*80)
        print("AWS PROFILE SELECTION")
        print("="*80)
        print("\nAvailable AWS profiles:")
        for i, profile in enumerate(profiles, 1):
            print(f"  {i}. {profile}")
        
        # Profile selection logic
        while True:
            try:
                if len(profiles) == 1:
                    print(f"\nOnly one profile available, using '{profiles[0]}'")
                    profile_name = profiles[0]
                    break
                else:
                    selection = input("\nEnter profile number or name (or press Enter for default): ").strip()
                    
                    if not selection:
                        profile_name = "default"
                        break
                    
                    if selection.isdigit() and 1 <= int(selection) <= len(profiles):
                        profile_name = profiles[int(selection) - 1]
                        break
                    elif selection in profiles:
                        profile_name = selection
                        break
                    else:
                        print("Invalid selection. Please try again.")
            except Exception as e:
                print(f"Error in profile selection: {e}")
        
        # Create session and validate
        try:
            self.session = boto3.Session(profile_name=profile_name)
            self.profile_name = profile_name
            sts = self.session.client('sts')
            
            identity = sts.get_caller_identity()
            self.account_id = identity['Account']
            iam_arn = identity['Arn']
            
            # Try to get account alias
            try:
                iam = self.session.client('iam')
                aliases = iam.list_account_aliases()
                if aliases['AccountAliases']:
                    self.account_alias = aliases['AccountAliases'][0]
            except:
                pass
            
            print("\n" + "="*80)
            print("AWS ACCOUNT INFORMATION")
            print("="*80)
            print(f"Profile:        {profile_name}")
            print(f"Account ID:     {self.account_id}")
            if self.account_alias:
                print(f"Account Alias:  {self.account_alias}")
            print(f"IAM ARN:        {iam_arn}")
            print("="*80)
            
            confirmation = input("\nIs this the correct account? (yes/no): ").lower()
            if confirmation not in ['y', 'yes']:
                print("\nAborting operation.")
                sys.exit(0)
            
            return True
            
        except Exception as e:
            print(f"\nError connecting to AWS with profile '{profile_name}': {e}")
            sys.exit(1)
    
    def setup_interactive(self):
        """Interactive setup for the analyzer"""
        print("\n" + "="*80)
        print("AWS NAT GATEWAY ANALYZER")
        print("="*80)
        print()
        
        # Get profile and account info first
        self.get_profile_and_account_info()
        
        print()
        
        # Region Selection
        print("="*80)
        print("REGION SELECTION")
        print("="*80)
        print("\nOptions:")
        print("  1. All regions (comprehensive but slower)")
        print("  2. Specific regions (faster)")
        print("  3. Current region only")
        
        choice = input("\nSelect an option [1-3] (default: 2): ").strip() or "2"
        
        if choice == "1":
            print("\nFetching all available regions...")
            self.regions = self._get_all_regions()
            print(f"✓ Will analyze {len(self.regions)} regions")
        elif choice == "3":
            current_region = self.session.region_name or 'us-east-1'
            self.regions = [current_region]
            print(f"✓ Will analyze current region: {current_region}")
        else:  # choice == "2" or default
            print("\nCommon regions:")
            common_regions = [
                'us-east-1', 'us-east-2', 'us-west-1', 'us-west-2',
                'eu-west-1', 'eu-central-1', 'ap-southeast-1', 'ap-northeast-1'
            ]
            for idx, region in enumerate(common_regions, 1):
                print(f"  {idx}. {region}")
            
            print("\nEnter region names separated by commas (e.g., us-east-1,us-west-2)")
            print("Or enter numbers from the list above (e.g., 1,4)")
            region_input = input("Regions: ").strip()
            
            if not region_input:
                self.regions = ['us-east-1']
                print("✓ Using default: us-east-1")
            else:
                # Check if input is numbers or region names
                if all(part.strip().isdigit() for part in region_input.split(',')):
                    # Numbers provided
                    indices = [int(x.strip()) - 1 for x in region_input.split(',')]
                    self.regions = [common_regions[i] for i in indices if 0 <= i < len(common_regions)]
                else:
                    # Region names provided
                    self.regions = [r.strip() for r in region_input.split(',')]
                
                print(f"✓ Will analyze {len(self.regions)} region(s): {', '.join(self.regions)}")
        
        print()
        
        # Output Options
        print("="*80)
        print("OUTPUT OPTIONS")
        print("="*80)
        print("\nOutput formats will be generated:")
        print("  ✓ Console output (always generated)")
        print("  ✓ Text file report")
        if EXCEL_AVAILABLE:
            print("  ✓ Excel workbook (single file with multiple tabs)")
        else:
            print("  ✗ Excel not available (install openpyxl for Excel support)")
        
        self.output_prefix = input("\nEnter output filename prefix (default: nat_gateway_analysis): ").strip() or "nat_gateway_analysis"
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.output_prefix = f"{self.output_prefix}_{timestamp}"
        
        print(f"✓ Output files will be saved with prefix: {self.output_prefix}")
        print()
        
        # Confirmation
        print("="*80)
        print("ANALYSIS CONFIGURATION SUMMARY")
        print("="*80)
        print(f"AWS Profile:     {self.profile_name}")
        print(f"AWS Account:     {self.account_id}" + (f" ({self.account_alias})" if self.account_alias else ""))
        print(f"Regions:         {', '.join(self.regions)}")
        print(f"Output Prefix:   {self.output_prefix}")
        print("="*80)
        print()
        
        confirm = input("Start analysis? (yes/no) [default: yes]: ").strip().lower() or 'yes'
        
        if confirm not in ['y', 'yes']:
            print("\nAnalysis cancelled.")
            sys.exit(0)
        
        print("\n" + "="*80)
        print("STARTING ANALYSIS...")
        print("="*80)
        print()
    
    def _get_all_regions(self):
        """Get all available AWS regions"""
        try:
            ec2 = self.session.client('ec2', region_name='us-east-1')
            regions = ec2.describe_regions()['Regions']
            return [region['RegionName'] for region in regions]
        except Exception as e:
            print(f"Error fetching regions: {str(e)}")
            return ['us-east-1']
    
    def analyze(self):
        """Main analysis function"""
        for region in self.regions:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Checking region: {region}")
            try:
                nat_gateways = self._get_nat_gateways(region)
                
                if nat_gateways:
                    print(f"  ✓ Found {len(nat_gateways)} NAT Gateway(s)")
                    
                    for nat_gw in nat_gateways:
                        nat_info = self._analyze_nat_gateway(region, nat_gw)
                        self.results.append(nat_info)
                        print(f"    - Analyzed {nat_gw['NatGatewayId']}")
                else:
                    print(f"  - No NAT Gateways found")
                    
            except Exception as e:
                print(f"  ✗ Error analyzing region {region}: {str(e)}")
            
            print()
        
        return self.results
    
    def _get_nat_gateways(self, region):
        """Get all NAT Gateways in a region"""
        ec2 = self.session.client('ec2', region_name=region)
        response = ec2.describe_nat_gateways()
        
        # Filter out deleted NAT Gateways
        return [nat for nat in response['NatGateways'] if nat['State'] != 'deleted']
    
    def _analyze_nat_gateway(self, region, nat_gw):
        """Analyze a single NAT Gateway and find resources using it"""
        ec2 = self.session.client('ec2', region_name=region)
        
        nat_id = nat_gw['NatGatewayId']
        vpc_id = nat_gw['VpcId']
        subnet_id = nat_gw['SubnetId']
        state = nat_gw['State']
        
        # Get Elastic IP
        elastic_ip = None
        for address in nat_gw.get('NatGatewayAddresses', []):
            if 'PublicIp' in address:
                elastic_ip = address['PublicIp']
                break
        
        # Get subnet details
        subnet_info = self._get_subnet_info(ec2, subnet_id)
        
        # Get VPC details
        vpc_info = self._get_vpc_info(ec2, vpc_id)
        
        # Get tags
        tags = {tag['Key']: tag['Value'] for tag in nat_gw.get('Tags', [])}
        nat_name = tags.get('Name', 'N/A')
        
        # Find route tables using this NAT Gateway
        route_tables = self._find_route_tables_using_nat(ec2, nat_id)
        
        # Find subnets associated with those route tables
        associated_subnets = self._find_associated_subnets(ec2, route_tables)
        
        # Find resources in those subnets
        resources = self._find_resources_in_subnets(ec2, region, associated_subnets)
        
        nat_info = {
            'nat_gateway_id': nat_id,
            'nat_gateway_name': nat_name,
            'region': region,
            'state': state,
            'vpc_id': vpc_id,
            'vpc_name': vpc_info.get('name', 'N/A'),
            'vpc_cidr': vpc_info.get('cidr', 'N/A'),
            'subnet_id': subnet_id,
            'subnet_name': subnet_info.get('name', 'N/A'),
            'subnet_cidr': subnet_info.get('cidr', 'N/A'),
            'availability_zone': subnet_info.get('az', 'N/A'),
            'elastic_ip': elastic_ip,
            'tags': tags,
            'route_tables': route_tables,
            'associated_subnets': associated_subnets,
            'resources': resources,
            'created_time': nat_gw.get('CreateTime', 'N/A')
        }
        
        return nat_info
    
    def _get_subnet_info(self, ec2, subnet_id):
        """Get subnet information"""
        try:
            response = ec2.describe_subnets(SubnetIds=[subnet_id])
            subnet = response['Subnets'][0]
            return {
                'cidr': subnet.get('CidrBlock'),
                'az': subnet.get('AvailabilityZone'),
                'name': next((tag['Value'] for tag in subnet.get('Tags', []) if tag['Key'] == 'Name'), 'N/A')
            }
        except Exception as e:
            return {'cidr': 'N/A', 'az': 'N/A', 'name': 'N/A'}
    
    def _get_vpc_info(self, ec2, vpc_id):
        """Get VPC information"""
        try:
            response = ec2.describe_vpcs(VpcIds=[vpc_id])
            vpc = response['Vpcs'][0]
            return {
                'cidr': vpc.get('CidrBlock'),
                'name': next((tag['Value'] for tag in vpc.get('Tags', []) if tag['Key'] == 'Name'), 'N/A')
            }
        except Exception as e:
            return {'cidr': 'N/A', 'name': 'N/A'}
    
    def _find_route_tables_using_nat(self, ec2, nat_id):
        """Find route tables that have routes to this NAT Gateway"""
        route_tables = []
        
        try:
            response = ec2.describe_route_tables()
            
            for rt in response['RouteTables']:
                for route in rt.get('Routes', []):
                    if route.get('NatGatewayId') == nat_id:
                        rt_name = next((tag['Value'] for tag in rt.get('Tags', []) if tag['Key'] == 'Name'), 'N/A')
                        route_tables.append({
                            'route_table_id': rt['RouteTableId'],
                            'route_table_name': rt_name,
                            'destination': route.get('DestinationCidrBlock', route.get('DestinationPrefixListId', 'N/A'))
                        })
                        break
        except Exception as e:
            pass
        
        return route_tables
    
    def _find_associated_subnets(self, ec2, route_tables):
        """Find subnets associated with the route tables"""
        subnets = []
        
        try:
            rt_ids = [rt['route_table_id'] for rt in route_tables]
            
            if not rt_ids:
                return subnets
            
            response = ec2.describe_route_tables(RouteTableIds=rt_ids)
            
            for rt in response['RouteTables']:
                for association in rt.get('Associations', []):
                    if 'SubnetId' in association:
                        subnet_id = association['SubnetId']
                        subnet_info = self._get_subnet_info(ec2, subnet_id)
                        subnets.append({
                            'subnet_id': subnet_id,
                            'subnet_name': subnet_info.get('name'),
                            'subnet_cidr': subnet_info.get('cidr'),
                            'availability_zone': subnet_info.get('az')
                        })
        except Exception as e:
            pass
        
        return subnets
    
    def _find_resources_in_subnets(self, ec2, region, subnets):
        """Find resources in the specified subnets"""
        resources = {
            'ec2_instances': [],
            'lambda_functions': [],
            'rds_instances': [],
            'ecs_tasks': []
        }
        
        if not subnets:
            return resources
        
        subnet_ids = [subnet['subnet_id'] for subnet in subnets]
        
        # Find EC2 instances
        try:
            response = ec2.describe_instances(
                Filters=[
                    {'Name': 'subnet-id', 'Values': subnet_ids},
                    {'Name': 'instance-state-name', 'Values': ['running', 'stopped', 'stopping', 'pending']}
                ]
            )
            
            for reservation in response['Reservations']:
                for instance in reservation['Instances']:
                    instance_name = next((tag['Value'] for tag in instance.get('Tags', []) if tag['Key'] == 'Name'), 'N/A')
                    resources['ec2_instances'].append({
                        'instance_id': instance['InstanceId'],
                        'instance_name': instance_name,
                        'instance_type': instance['InstanceType'],
                        'state': instance['State']['Name'],
                        'private_ip': instance.get('PrivateIpAddress', 'N/A'),
                        'subnet_id': instance['SubnetId']
                    })
        except Exception as e:
            pass
        
        # Find Lambda functions
        try:
            lambda_client = self.session.client('lambda', region_name=region)
            paginator = lambda_client.get_paginator('list_functions')
            
            for page in paginator.paginate():
                for function in page['Functions']:
                    vpc_config = function.get('VpcConfig', {})
                    function_subnets = vpc_config.get('SubnetIds', [])
                    
                    if any(subnet_id in function_subnets for subnet_id in subnet_ids):
                        resources['lambda_functions'].append({
                            'function_name': function['FunctionName'],
                            'runtime': function['Runtime'],
                            'subnets': function_subnets
                        })
        except Exception as e:
            pass
        
        # Find RDS instances
        try:
            rds_client = self.session.client('rds', region_name=region)
            response = rds_client.describe_db_instances()
            
            for db in response['DBInstances']:
                db_subnets = [subnet['SubnetIdentifier'] for subnet in db.get('DBSubnetGroup', {}).get('Subnets', [])]
                
                if any(subnet_id in db_subnets for subnet_id in subnet_ids):
                    resources['rds_instances'].append({
                        'db_instance_id': db['DBInstanceIdentifier'],
                        'engine': db['Engine'],
                        'status': db['DBInstanceStatus'],
                        'subnets': db_subnets
                    })
        except Exception as e:
            pass
        
        return resources
    
    def print_console_report(self):
        """Print analysis results to console"""
        print("\n")
        print("=" * 100)
        print("NAT GATEWAY ANALYSIS REPORT")
        print("=" * 100)
        print(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"AWS Profile: {self.profile_name}")
        print(f"AWS Account: {self.account_id}" + (f" ({self.account_alias})" if self.account_alias else ""))
        print(f"Regions Analyzed: {', '.join(self.regions)}")
        print(f"Total NAT Gateways Found: {len(self.results)}")
        print("=" * 100)
        print()
        
        if not self.results:
            print("No NAT Gateways found in the analyzed regions.")
            return
        
        for idx, nat in enumerate(self.results, 1):
            print(f"\n{'#' * 100}")
            print(f"NAT GATEWAY #{idx}: {nat['nat_gateway_id']}")
            print(f"{'#' * 100}")
            print()
            
            # Basic Information
            print("BASIC INFORMATION")
            print("-" * 100)
            print(f"  NAT Gateway ID:        {nat['nat_gateway_id']}")
            print(f"  Name:                  {nat['nat_gateway_name']}")
            print(f"  Region:                {nat['region']}")
            print(f"  State:                 {nat['state']}")
            print(f"  Elastic IP:            {nat['elastic_ip']}")
            print(f"  Created:               {nat['created_time']}")
            print()
            
            # Location Information
            print("LOCATION")
            print("-" * 100)
            print(f"  VPC ID:                {nat['vpc_id']}")
            print(f"  VPC Name:              {nat['vpc_name']}")
            print(f"  VPC CIDR:              {nat['vpc_cidr']}")
            print(f"  Subnet ID:             {nat['subnet_id']}")
            print(f"  Subnet Name:           {nat['subnet_name']}")
            print(f"  Subnet CIDR:           {nat['subnet_cidr']}")
            print(f"  Availability Zone:     {nat['availability_zone']}")
            print()
            
            # Tags
            if nat['tags']:
                print("TAGS")
                print("-" * 100)
                for key, value in nat['tags'].items():
                    print(f"  {key:20} = {value}")
                print()
            
            # Route Tables
            print(f"ROUTE TABLES USING THIS NAT GATEWAY ({len(nat['route_tables'])})")
            print("-" * 100)
            if nat['route_tables']:
                for rt in nat['route_tables']:
                    print(f"  Route Table ID:   {rt['route_table_id']}")
                    print(f"  Name:             {rt['route_table_name']}")
                    print(f"  Destination:      {rt['destination']}")
                    print()
            else:
                print("  No route tables found using this NAT Gateway")
                print()
            
            # Associated Subnets
            print(f"ASSOCIATED SUBNETS ({len(nat['associated_subnets'])})")
            print("-" * 100)
            if nat['associated_subnets']:
                for subnet in nat['associated_subnets']:
                    print(f"  Subnet ID:        {subnet['subnet_id']}")
                    print(f"  Name:             {subnet['subnet_name']}")
                    print(f"  CIDR:             {subnet['subnet_cidr']}")
                    print(f"  AZ:               {subnet['availability_zone']}")
                    print()
            else:
                print("  No associated subnets found")
                print()
            
            # Resources
            resources = nat['resources']
            total_resources = (len(resources['ec2_instances']) + 
                             len(resources['lambda_functions']) + 
                             len(resources['rds_instances']))
            
            print(f"RESOURCES USING THIS NAT GATEWAY ({total_resources} total)")
            print("-" * 100)
            
            if resources['ec2_instances']:
                print(f"\n  EC2 INSTANCES ({len(resources['ec2_instances'])})")
                print("  " + "-" * 96)
                for instance in resources['ec2_instances']:
                    print(f"    Instance ID:    {instance['instance_id']}")
                    print(f"    Name:           {instance['instance_name']}")
                    print(f"    Type:           {instance['instance_type']}")
                    print(f"    State:          {instance['state']}")
                    print(f"    Private IP:     {instance['private_ip']}")
                    print(f"    Subnet:         {instance['subnet_id']}")
                    print()
            
            if resources['lambda_functions']:
                print(f"\n  LAMBDA FUNCTIONS ({len(resources['lambda_functions'])})")
                print("  " + "-" * 96)
                for func in resources['lambda_functions']:
                    print(f"    Function Name:  {func['function_name']}")
                    print(f"    Runtime:        {func['runtime']}")
                    print(f"    Subnets:        {', '.join(func['subnets'])}")
                    print()
            
            if resources['rds_instances']:
                print(f"\n  RDS INSTANCES ({len(resources['rds_instances'])})")
                print("  " + "-" * 96)
                for db in resources['rds_instances']:
                    print(f"    DB Instance ID: {db['db_instance_id']}")
                    print(f"    Engine:         {db['engine']}")
                    print(f"    Status:         {db['status']}")
                    print(f"    Subnets:        {', '.join(db['subnets'])}")
                    print()
            
            if total_resources == 0:
                print("  No resources found using this NAT Gateway")
                print()
        
        print("\n" + "=" * 100)
        print("END OF REPORT")
        print("=" * 100)
    
    def export_to_text(self):
        """Export results to text file"""
        filename = f"{self.output_prefix}.txt"
        
        with open(filename, 'w') as f:
            f.write("=" * 100 + "\n")
            f.write("NAT GATEWAY ANALYSIS REPORT\n")
            f.write("=" * 100 + "\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"AWS Profile: {self.profile_name}\n")
            f.write(f"AWS Account: {self.account_id}" + (f" ({self.account_alias})" if self.account_alias else "") + "\n")
            f.write(f"Regions Analyzed: {', '.join(self.regions)}\n")
            f.write(f"Total NAT Gateways Found: {len(self.results)}\n")
            f.write("=" * 100 + "\n\n")
            
            if not self.results:
                f.write("No NAT Gateways found in the analyzed regions.\n")
                return filename
            
            for idx, nat in enumerate(self.results, 1):
                f.write("\n" + "#" * 100 + "\n")
                f.write(f"NAT GATEWAY #{idx}: {nat['nat_gateway_id']}\n")
                f.write("#" * 100 + "\n\n")
                
                # Basic Information
                f.write("BASIC INFORMATION\n")
                f.write("-" * 100 + "\n")
                f.write(f"  NAT Gateway ID:        {nat['nat_gateway_id']}\n")
                f.write(f"  Name:                  {nat['nat_gateway_name']}\n")
                f.write(f"  Region:                {nat['region']}\n")
                f.write(f"  State:                 {nat['state']}\n")
                f.write(f"  Elastic IP:            {nat['elastic_ip']}\n")
                f.write(f"  Created:               {nat['created_time']}\n\n")
                
                # Location Information
                f.write("LOCATION\n")
                f.write("-" * 100 + "\n")
                f.write(f"  VPC ID:                {nat['vpc_id']}\n")
                f.write(f"  VPC Name:              {nat['vpc_name']}\n")
                f.write(f"  VPC CIDR:              {nat['vpc_cidr']}\n")
                f.write(f"  Subnet ID:             {nat['subnet_id']}\n")
                f.write(f"  Subnet Name:           {nat['subnet_name']}\n")
                f.write(f"  Subnet CIDR:           {nat['subnet_cidr']}\n")
                f.write(f"  Availability Zone:     {nat['availability_zone']}\n\n")
                
                # Tags
                if nat['tags']:
                    f.write("TAGS\n")
                    f.write("-" * 100 + "\n")
                    for key, value in nat['tags'].items():
                        f.write(f"  {key:20} = {value}\n")
                    f.write("\n")
                
                # Route Tables
                f.write(f"ROUTE TABLES USING THIS NAT GATEWAY ({len(nat['route_tables'])})\n")
                f.write("-" * 100 + "\n")
                if nat['route_tables']:
                    for rt in nat['route_tables']:
                        f.write(f"  Route Table ID:   {rt['route_table_id']}\n")
                        f.write(f"  Name:             {rt['route_table_name']}\n")
                        f.write(f"  Destination:      {rt['destination']}\n\n")
                else:
                    f.write("  No route tables found using this NAT Gateway\n\n")
                
                # Associated Subnets
                f.write(f"ASSOCIATED SUBNETS ({len(nat['associated_subnets'])})\n")
                f.write("-" * 100 + "\n")
                if nat['associated_subnets']:
                    for subnet in nat['associated_subnets']:
                        f.write(f"  Subnet ID:        {subnet['subnet_id']}\n")
                        f.write(f"  Name:             {subnet['subnet_name']}\n")
                        f.write(f"  CIDR:             {subnet['subnet_cidr']}\n")
                        f.write(f"  AZ:               {subnet['availability_zone']}\n\n")
                else:
                    f.write("  No associated subnets found\n\n")
                
                # Resources
                resources = nat['resources']
                total_resources = (len(resources['ec2_instances']) + 
                                 len(resources['lambda_functions']) + 
                                 len(resources['rds_instances']))
                
                f.write(f"RESOURCES USING THIS NAT GATEWAY ({total_resources} total)\n")
                f.write("-" * 100 + "\n")
                
                if resources['ec2_instances']:
                    f.write(f"\n  EC2 INSTANCES ({len(resources['ec2_instances'])})\n")
                    f.write("  " + "-" * 96 + "\n")
                    for instance in resources['ec2_instances']:
                        f.write(f"    Instance ID:    {instance['instance_id']}\n")
                        f.write(f"    Name:           {instance['instance_name']}\n")
                        f.write(f"    Type:           {instance['instance_type']}\n")
                        f.write(f"    State:          {instance['state']}\n")
                        f.write(f"    Private IP:     {instance['private_ip']}\n")
                        f.write(f"    Subnet:         {instance['subnet_id']}\n\n")
                
                if resources['lambda_functions']:
                    f.write(f"\n  LAMBDA FUNCTIONS ({len(resources['lambda_functions'])})\n")
                    f.write("  " + "-" * 96 + "\n")
                    for func in resources['lambda_functions']:
                        f.write(f"    Function Name:  {func['function_name']}\n")
                        f.write(f"    Runtime:        {func['runtime']}\n")
                        f.write(f"    Subnets:        {', '.join(func['subnets'])}\n\n")
                
                if resources['rds_instances']:
                    f.write(f"\n  RDS INSTANCES ({len(resources['rds_instances'])})\n")
                    f.write("  " + "-" * 96 + "\n")
                    for db in resources['rds_instances']:
                        f.write(f"    DB Instance ID: {db['db_instance_id']}\n")
                        f.write(f"    Engine:         {db['engine']}\n")
                        f.write(f"    Status:         {db['status']}\n")
                        f.write(f"    Subnets:        {', '.join(db['subnets'])}\n\n")
                
                if total_resources == 0:
                    f.write("  No resources found using this NAT Gateway\n\n")
            
            f.write("\n" + "=" * 100 + "\n")
            f.write("END OF REPORT\n")
            f.write("=" * 100 + "\n")
        
        return filename
    
    def export_to_excel(self):
        """Export results to Excel workbook with multiple sheets"""
        if not EXCEL_AVAILABLE:
            print("Warning: openpyxl not installed. Skipping Excel export.")
            return None
        
        filename = f"{self.output_prefix}.xlsx"
        wb = Workbook()
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        section_header_font = Font(bold=True, size=12, color="366092")
        bold_font = Font(bold=True)
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Helper function to style headers
        def style_header_row(ws, row=1):
            for cell in ws[row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_style
        
        # Helper function to auto-size columns
        def auto_size_columns(ws):
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 80)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Report Header
        ws_summary.append(['NAT Gateway Analysis Report'])
        ws_summary['A1'].font = Font(bold=True, size=16, color="366092")
        ws_summary.append([])
        
        # Analysis Information
        ws_summary.append(['ANALYSIS INFORMATION'])
        ws_summary['A3'].font = section_header_font
        ws_summary.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws_summary.append(['AWS Profile:', self.profile_name])
        ws_summary.append(['AWS Account:', f"{self.account_id}" + (f" ({self.account_alias})" if self.account_alias else "")])
        ws_summary.append(['Regions Analyzed:', ', '.join(self.regions)])
        ws_summary.append([])
        
        # Results Summary
        ws_summary.append(['RESULTS SUMMARY'])
        ws_summary['A9'].font = section_header_font
        ws_summary.append(['Total NAT Gateways Found:', len(self.results)])
        
        # Count totals
        total_route_tables = sum(len(nat['route_tables']) for nat in self.results)
        total_subnets = sum(len(nat['associated_subnets']) for nat in self.results)
        total_ec2 = sum(len(nat['resources']['ec2_instances']) for nat in self.results)
        total_lambda = sum(len(nat['resources']['lambda_functions']) for nat in self.results)
        total_rds = sum(len(nat['resources']['rds_instances']) for nat in self.results)
        
        ws_summary.append(['Total Route Tables:', total_route_tables])
        ws_summary.append(['Total Associated Subnets:', total_subnets])
        ws_summary.append(['Total EC2 Instances:', total_ec2])
        ws_summary.append(['Total Lambda Functions:', total_lambda])
        ws_summary.append(['Total RDS Instances:', total_rds])
        ws_summary.append(['Total Resources:', total_ec2 + total_lambda + total_rds])
        ws_summary.append([])
        
        # Tab Descriptions
        current_row = ws_summary.max_row + 1
        ws_summary.append(['WORKBOOK TABS GUIDE'])
        ws_summary[f'A{current_row}'].font = section_header_font
        ws_summary.append([])
        
        # Create a styled table for tab descriptions
        tab_descriptions = [
            {
                'tab': 'Summary',
                'description': 'Overview of the analysis report',
                'purpose': 'Provides high-level information about the analysis including account details, regions analyzed, total counts, and descriptions of all tabs in this workbook.',
                'contains': 'Report metadata, analysis parameters, summary statistics, and this guide'
            },
            {
                'tab': 'NAT Gateways',
                'description': 'Complete list of all NAT Gateways',
                'purpose': 'Main inventory of NAT Gateways with their configuration details, location information, and resource counts. Use this tab to get a complete overview of all NAT Gateways in the analyzed regions.',
                'contains': 'NAT Gateway ID, Name, Region, State, Elastic IP, VPC details, Subnet details, Availability Zone, resource counts, and creation time'
            },
            {
                'tab': 'Route Tables',
                'description': 'Route tables that route traffic through NAT Gateways',
                'purpose': 'Identifies which route tables are configured to send traffic through each NAT Gateway. This helps understand the routing configuration and traffic flow patterns.',
                'contains': 'Route Table ID, Route Table Name, associated NAT Gateway, destination CIDR blocks'
            },
            {
                'tab': 'Associated Subnets',
                'description': 'Subnets that use NAT Gateways for outbound traffic',
                'purpose': 'Lists all subnets whose route tables direct traffic through NAT Gateways. These are typically private subnets that need internet access for outbound connections.',
                'contains': 'Subnet ID, Subnet Name, CIDR blocks, Availability Zones, associated NAT Gateway'
            },
            {
                'tab': 'EC2 Instances',
                'description': 'EC2 instances using NAT Gateways',
                'purpose': 'Identifies all EC2 instances located in subnets that route outbound traffic through NAT Gateways. Useful for understanding which workloads depend on NAT Gateway connectivity.',
                'contains': 'Instance ID, Instance Name, Instance Type, State, Private IP, Subnet, associated NAT Gateway'
            },
            {
                'tab': 'Lambda Functions',
                'description': 'Lambda functions using NAT Gateways',
                'purpose': 'Lists Lambda functions deployed in VPC subnets that use NAT Gateways for external connectivity. Important for identifying serverless workloads with external dependencies.',
                'contains': 'Function Name, Runtime, VPC Subnets, associated NAT Gateway'
            },
            {
                'tab': 'RDS Instances',
                'description': 'RDS database instances using NAT Gateways',
                'purpose': 'Identifies RDS instances in subnets that may route traffic through NAT Gateways. Helps understand database connectivity patterns and potential external dependencies.',
                'contains': 'DB Instance ID, Engine type, Status, DB Subnets, associated NAT Gateway'
            }
        ]
        
        # Add header row for tab descriptions
        ws_summary.append(['Tab Name', 'Description', 'Purpose', 'What It Contains'])
        header_row = ws_summary.max_row
        
        for col in range(1, 5):
            cell = ws_summary.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = border_style
        
        # Add tab description rows
        for tab_info in tab_descriptions:
            ws_summary.append([
                tab_info['tab'],
                tab_info['description'],
                tab_info['purpose'],
                tab_info['contains']
            ])
            row = ws_summary.max_row
            
            # Style the row
            for col in range(1, 5):
                cell = ws_summary.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = border_style
                
                # Bold the tab name
                if col == 1:
                    cell.font = bold_font
        
        # Set column widths for summary sheet
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 35
        ws_summary.column_dimensions['C'].width = 50
        ws_summary.column_dimensions['D'].width = 50
        
        # Add notes section
        ws_summary.append([])
        ws_summary.append([])
        notes_row = ws_summary.max_row + 1
        ws_summary.append(['NOTES'])
        ws_summary[f'A{notes_row}'].font = section_header_font
        ws_summary.append(['• NAT Gateways in "deleted" state are excluded from this analysis'])
        ws_summary.append(['• Resources are identified based on subnet associations with route tables'])
        ws_summary.append(['• A resource "uses" a NAT Gateway if its subnet has a route table pointing to that NAT Gateway'])
        ws_summary.append(['• Each tab can be filtered and sorted independently for detailed analysis'])
        ws_summary.append(['• For questions or issues, contact your cloud infrastructure team'])
        
        # Sheet 2: NAT Gateways
        ws_nat = wb.create_sheet("NAT Gateways")
        ws_nat.append([
            'NAT Gateway ID', 'Name', 'Region', 'State', 'Elastic IP',
            'VPC ID', 'VPC Name', 'VPC CIDR',
            'Subnet ID', 'Subnet Name', 'Subnet CIDR', 'Availability Zone',
            'Route Tables Count', 'Associated Subnets Count',
            'EC2 Instances Count', 'Lambda Functions Count', 'RDS Instances Count',
            'Total Resources Count', 'Created Time'
        ])
        style_header_row(ws_nat)
        
        for nat in self.results:
            resources = nat['resources']
            ws_nat.append([
                nat['nat_gateway_id'],
                nat['nat_gateway_name'],
                nat['region'],
                nat['state'],
                nat['elastic_ip'],
                nat['vpc_id'],
                nat['vpc_name'],
                nat['vpc_cidr'],
                nat['subnet_id'],
                nat['subnet_name'],
                nat['subnet_cidr'],
                nat['availability_zone'],
                len(nat['route_tables']),
                len(nat['associated_subnets']),
                len(resources['ec2_instances']),
                len(resources['lambda_functions']),
                len(resources['rds_instances']),
                len(resources['ec2_instances']) + len(resources['lambda_functions']) + len(resources['rds_instances']),
                str(nat['created_time'])
            ])
        
        auto_size_columns(ws_nat)
        
        # Sheet 3: Route Tables
        ws_rt = wb.create_sheet("Route Tables")
        ws_rt.append([
            'NAT Gateway ID', 'NAT Gateway Name', 'Region',
            'Route Table ID', 'Route Table Name', 'Destination'
        ])
        style_header_row(ws_rt)
        
        for nat in self.results:
            for rt in nat['route_tables']:
                ws_rt.append([
                    nat['nat_gateway_id'],
                    nat['nat_gateway_name'],
                    nat['region'],
                    rt['route_table_id'],
                    rt['route_table_name'],
                    rt['destination']
                ])
        
        auto_size_columns(ws_rt)
        
        # Sheet 4: Associated Subnets
        ws_subnets = wb.create_sheet("Associated Subnets")
        ws_subnets.append([
            'NAT Gateway ID', 'NAT Gateway Name', 'Region',
            'Subnet ID', 'Subnet Name', 'Subnet CIDR', 'Availability Zone'
        ])
        style_header_row(ws_subnets)
        
        for nat in self.results:
            for subnet in nat['associated_subnets']:
                ws_subnets.append([
                    nat['nat_gateway_id'],
                    nat['nat_gateway_name'],
                    nat['region'],
                    subnet['subnet_id'],
                    subnet['subnet_name'],
                    subnet['subnet_cidr'],
                    subnet['availability_zone']
                ])
        
        auto_size_columns(ws_subnets)
        
        # Sheet 5: EC2 Instances
        ws_ec2 = wb.create_sheet("EC2 Instances")
        ws_ec2.append([
            'NAT Gateway ID', 'NAT Gateway Name', 'Region',
            'Instance ID', 'Instance Name', 'Instance Type',
            'State', 'Private IP', 'Subnet ID'
        ])
        style_header_row(ws_ec2)
        
        for nat in self.results:
            for instance in nat['resources']['ec2_instances']:
                ws_ec2.append([
                    nat['nat_gateway_id'],
                    nat['nat_gateway_name'],
                    nat['region'],
                    instance['instance_id'],
                    instance['instance_name'],
                    instance['instance_type'],
                    instance['state'],
                    instance['private_ip'],
                    instance['subnet_id']
                ])
        
        auto_size_columns(ws_ec2)
        
        # Sheet 6: Lambda Functions
        ws_lambda = wb.create_sheet("Lambda Functions")
        ws_lambda.append([
            'NAT Gateway ID', 'NAT Gateway Name', 'Region',
            'Function Name', 'Runtime', 'Subnets'
        ])
        style_header_row(ws_lambda)
        
        for nat in self.results:
            for func in nat['resources']['lambda_functions']:
                ws_lambda.append([
                    nat['nat_gateway_id'],
                    nat['nat_gateway_name'],
                    nat['region'],
                    func['function_name'],
                    func['runtime'],
                    ', '.join(func['subnets'])
                ])
        
        auto_size_columns(ws_lambda)
        
        # Sheet 7: RDS Instances
        ws_rds = wb.create_sheet("RDS Instances")
        ws_rds.append([
            'NAT Gateway ID', 'NAT Gateway Name', 'Region',
            'DB Instance ID', 'Engine', 'Status', 'Subnets'
        ])
        style_header_row(ws_rds)
        
        for nat in self.results:
            for db in nat['resources']['rds_instances']:
                ws_rds.append([
                    nat['nat_gateway_id'],
                    nat['nat_gateway_name'],
                    nat['region'],
                    db['db_instance_id'],
                    db['engine'],
                    db['status'],
                    ', '.join(db['subnets'])
                ])
        
        auto_size_columns(ws_rds)
        
        # Save workbook
        wb.save(filename)
        
        return filename


def main():
    """Main execution function"""
    analyzer = NATGatewayAnalyzer()
    
    # Interactive setup
    analyzer.setup_interactive()
    
    # Run analysis
    results = analyzer.analyze()
    
    print("\n" + "="*80)
    print("ANALYSIS COMPLETE")
    print("="*80)
    print(f"Total NAT Gateways Found: {len(results)}")
    print()
    
    # Generate outputs
    print("Generating outputs...")
    print()
    
    # Console output
    analyzer.print_console_report()
    
    # Text file
    text_file = analyzer.export_to_text()
    print(f"\n✓ Text report saved: {text_file}")
    
    # Excel file (if available)
    if EXCEL_AVAILABLE:
        excel_file = analyzer.export_to_excel()
        if excel_file:
            print(f"✓ Excel workbook saved: {excel_file}")
            print(f"  Contains 7 tabs:")
            print(f"    - Summary (with detailed tab guide)")
            print(f"    - NAT Gateways")
            print(f"    - Route Tables")
            print(f"    - Associated Subnets")
            print(f"    - EC2 Instances")
            print(f"    - Lambda Functions")
            print(f"    - RDS Instances")
    else:
        print("\n✗ Excel export skipped (openpyxl not installed)")
        print("  Install with: pip install openpyxl")
    
    print()
    print("="*80)
    print("ALL OUTPUTS GENERATED SUCCESSFULLY")
    print("="*80)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nAnalysis interrupted by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\n\nError: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
