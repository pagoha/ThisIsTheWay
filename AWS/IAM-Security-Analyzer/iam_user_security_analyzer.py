#!/usr/bin/env python3
"""
AWS IAM User Security Analyzer
Comprehensive analysis of IAM users, access keys, policies, and MFA status
"""

import boto3
import botocore.session
import sys
import time
import csv
import os
import logging
import concurrent.futures
from threading import Lock
from datetime import datetime, timezone, timedelta
from botocore.exceptions import ClientError

# Try to import openpyxl for Excel support, fall back to CSV if not available
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will not be available.")
    print("Install with: pip install openpyxl")

# Default configuration
DEFAULT_CONFIG = {
    'days_threshold': 200,
    'max_workers': 5,
    'retry_attempts': 3,
    'output_formats': ['txt', 'excel'],
    'exclude_patterns': [],
    'check_mfa': True,
    'analyze_policies': True,
    'analyze_access_keys': True,
    'dry_run': False
}

class IAMUserSecurityAnalyzer:
    def __init__(self, config=None):
        self.config = config or DEFAULT_CONFIG.copy()
        self.setup_logging()
        self.results_lock = Lock()
        
    def setup_logging(self):
        """Setup logging configuration"""
        log_format = '%(asctime)s - %(levelname)s - %(message)s'
        logging.basicConfig(
            level=logging.INFO,
            format=log_format,
            handlers=[
                logging.FileHandler('iam_user_security_analyzer.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
    
    def get_user_configuration(self):
        """Prompt user for configuration options"""
        print("\n" + "="*60)
        print("CONFIGURATION OPTIONS".center(60))
        print("="*60)
        
        # Inactivity threshold
        threshold_input = input(f"Enter inactivity threshold in days (default {self.config['days_threshold']}): ")
        if threshold_input and threshold_input.isdigit():
            self.config['days_threshold'] = int(threshold_input)
        
        # Max workers
        workers_input = input(f"Enter max worker threads for parallel processing (default {self.config['max_workers']}): ")
        if workers_input and workers_input.isdigit() and int(workers_input) > 0:
            self.config['max_workers'] = int(workers_input)
        
        # Analyze policies
        policies_input = input("Analyze attached policies? (y/n, default y): ").lower()
        if policies_input in ['n', 'no']:
            self.config['analyze_policies'] = False
        
        # Analyze access keys
        keys_input = input("Perform detailed access key analysis? (y/n, default y): ").lower()
        if keys_input in ['n', 'no']:
            self.config['analyze_access_keys'] = False
        
        # Output formats
        print("\nOutput format options:")
        print("1. Text report only")
        if EXCEL_AVAILABLE:
            print("2. Excel workbook only (multiple tabs)")
            print("3. Both text and Excel (default)")
        else:
            print("2. CSV files only")
            print("3. Both text and CSV (default)")
        format_choice = input("Choose output format (1-3, default 3): ")
        
        if format_choice == '1':
            self.config['output_formats'] = ['txt']
        elif format_choice == '2':
            self.config['output_formats'] = ['excel' if EXCEL_AVAILABLE else 'csv']
        else:
            self.config['output_formats'] = ['txt', 'excel' if EXCEL_AVAILABLE else 'csv']
        
        # Exclude patterns
        exclude_input = input("Enter exclude patterns for usernames (comma-separated, optional): ")
        if exclude_input:
            self.config['exclude_patterns'] = [p.strip() for p in exclude_input.split(',')]
        
        # Dry run
        dry_run_input = input("Dry run mode (skip confirmations)? (y/n, default n): ").lower()
        if dry_run_input in ['y', 'yes']:
            self.config['dry_run'] = True
        
        print(f"\nConfiguration Summary:")
        print(f"  - Inactivity threshold: {self.config['days_threshold']} days")
        print(f"  - Max worker threads: {self.config['max_workers']}")
        print(f"  - Check MFA status: {self.config['check_mfa']} (always enabled)")
        print(f"  - Analyze policies: {self.config['analyze_policies']}")
        print(f"  - Analyze access keys: {self.config['analyze_access_keys']}")
        print(f"  - Output formats: {', '.join(self.config['output_formats'])}")
        print(f"  - Exclude patterns: {self.config['exclude_patterns']}")
        print(f"  - Dry run mode: {self.config['dry_run']}")
    
    def make_api_call_with_retry(self, func, *args, **kwargs):
        """Make API calls with exponential backoff retry"""
        max_retries = self.config['retry_attempts']
        
        for attempt in range(max_retries):
            try:
                return func(*args, **kwargs)
            except ClientError as e:
                error_code = e.response['Error']['Code']
                if error_code in ['Throttling', 'RequestLimitExceeded']:
                    wait_time = (2 ** attempt) + (time.time() % 1)
                    self.logger.warning(f"Rate limited, waiting {wait_time:.2f} seconds...")
                    time.sleep(wait_time)
                    continue
                else:
                    self.logger.error(f"API call failed: {e}")
                    raise
            except Exception as e:
                self.logger.error(f"Unexpected error in API call: {e}")
                if attempt == max_retries - 1:
                    raise
                time.sleep(1)
        
        raise Exception(f"Max retries ({max_retries}) exceeded")
    
    def get_profile_and_account_info(self):
        """Prompt for AWS profile and confirm account details"""
        session = botocore.session.Session()
        profiles = session.available_profiles
        
        if not profiles:
            self.logger.error("No AWS profiles found. Please configure AWS CLI first.")
            sys.exit(1)
        
        print("\nAvailable AWS profiles:")
        for i, profile in enumerate(profiles, 1):
            print(f"{i}. {profile}")
        
        # Profile selection logic
        while True:
            try:
                if len(profiles) == 1:
                    print(f"\nOnly one profile available, using '{profiles[0]}'")
                    profile_name = profiles[0]
                    break
                else:
                    selection = input("\nEnter profile number or name (or press Enter for default): ")
                    
                    if not selection:
                        profile_name = "default"
                        break
                    
                    if selection.isdigit() and 1 <= int(selection) <= len(profiles):
                        profile_name = profiles[int(selection) - 1]
                        break
                    elif selection in profiles:
                        profile_name = selection
                        break
                    else:
                        print("Invalid selection. Please try again.")
            except Exception as e:
                self.logger.error(f"Error in profile selection: {e}")
        
        # Create session and validate
        try:
            session = boto3.Session(profile_name=profile_name)
            sts = session.client('sts')
            
            identity = self.make_api_call_with_retry(sts.get_caller_identity)
            account_id = identity['Account']
            iam_arn = identity['Arn']
            
            print("\nAWS Account Information:")
            print(f"Profile: {profile_name}")
            print(f"Account ID: {account_id}")
            print(f"IAM ARN: {iam_arn}")
            
            if not self.config['dry_run']:
                confirmation = input("\nIs this the correct account? (yes/no): ").lower()
                if confirmation not in ['y', 'yes']:
                    print("Aborting operation.")
                    sys.exit(0)
            
            return session, account_id, profile_name
            
        except Exception as e:
            self.logger.error(f"Error connecting to AWS with profile '{profile_name}': {e}")
            sys.exit(1)
    
    def check_required_permissions(self, iam):
        """Validate that we have the required IAM permissions"""
        required_actions = [
            'iam:ListUsers',
            'iam:ListAccessKeys',
            'iam:GetAccessKeyLastUsed',
            'iam:ListMFADevices',
            'iam:ListAttachedUserPolicies',
            'iam:ListUserPolicies'
        ]
        
        try:
            # Try a simple list operation to check permissions
            self.make_api_call_with_retry(iam.list_users, MaxItems=1)
            self.logger.info("IAM permissions validated successfully")
            return True
        except ClientError as e:
            if e.response['Error']['Code'] == 'AccessDenied':
                self.logger.error("Insufficient IAM permissions. Required actions:")
                for action in required_actions:
                    self.logger.error(f"  - {action}")
                return False
            raise
    
    def analyze_policies(self, iam, username):
        """Analyze attached policies for users"""
        policies = []
        risky_policies = []
        
        try:
            # Get user policies
            attached = self.make_api_call_with_retry(
                iam.list_attached_user_policies, UserName=username
            )
            inline = self.make_api_call_with_retry(
                iam.list_user_policies, UserName=username
            )
            
            for policy in attached.get('AttachedPolicies', []):
                policy['Type'] = 'Managed'
                policies.append(policy)
                # Check for risky policies
                if any(keyword in policy['PolicyName'].lower() 
                       for keyword in ['admin', 'full', '*', 'poweruser']):
                    risky_policies.append(policy['PolicyName'])
            
            for policy_name in inline.get('PolicyNames', []):
                policy_info = {'PolicyName': policy_name, 'Type': 'Inline'}
                policies.append(policy_info)
                if any(keyword in policy_name.lower() 
                       for keyword in ['admin', 'full', '*', 'poweruser']):
                    risky_policies.append(policy_name)
        
        except Exception as e:
            self.logger.warning(f"Error analyzing policies: {e}")
        
        return policies, risky_policies
    
    def check_mfa_status(self, iam, username):
        """Check MFA status for a user"""
        try:
            mfa_devices = self.make_api_call_with_retry(
                iam.list_mfa_devices, UserName=username
            )
            return len(mfa_devices.get('MFADevices', [])) > 0
        except Exception as e:
            self.logger.warning(f"Error checking MFA for user {username}: {e}")
            return None
    
    def filter_by_tags(self, iam, resource_name):
        """Check resource tags for exclusion patterns"""
        try:
            tags_response = self.make_api_call_with_retry(
                iam.list_user_tags, UserName=resource_name
            )
            
            tag_dict = {tag['Key']: tag['Value'] for tag in tags_response.get('Tags', [])}
            
            # Skip if marked as system/service account
            if tag_dict.get('Type', '').lower() in ['system', 'service', 'automation']:
                return True
            
            # Check against exclude patterns
            for pattern in self.config['exclude_patterns']:
                if pattern and pattern.lower() in resource_name.lower():
                    return True
            
            return False
        except:
            return False
    
    def analyze_access_keys_detailed(self, iam, username, today, days_threshold):
        """Detailed analysis of access keys for a user"""
        access_keys_info = {
            'total_keys': 0,
            'active_keys': 0,
            'old_unused_keys': [],
            'old_used_keys': [],
            'latest_key_usage': None,
            'all_keys': []
        }
        
        try:
            keys_response = self.make_api_call_with_retry(
                iam.list_access_keys, UserName=username
            )
            access_keys = keys_response.get('AccessKeyMetadata', [])
            access_keys_info['total_keys'] = len(access_keys)
            
            for key in access_keys:
                access_key_id = key['AccessKeyId']
                key_status = key['Status']
                key_created_date = key['CreateDate']
                key_age_days = (today - key_created_date).days
                
                if key_status == 'Active':
                    access_keys_info['active_keys'] += 1
                
                # Get key last usage
                last_used_response = self.make_api_call_with_retry(
                    iam.get_access_key_last_used, AccessKeyId=access_key_id
                )
                last_used_info = last_used_response.get('AccessKeyLastUsed', {})
                last_used_date = last_used_info.get('LastUsedDate')
                
                key_info = {
                    'AccessKeyId': access_key_id,
                    'Status': key_status,
                    'Created': key_created_date,
                    'AgeInDays': key_age_days,
                    'LastUsed': last_used_date,
                    'LastUsedService': last_used_info.get('ServiceName', 'N/A'),
                    'LastUsedRegion': last_used_info.get('Region', 'N/A')
                }
                
                if last_used_date:
                    days_since_last_use = (today - last_used_date).days
                    key_info['DaysSinceLastUse'] = days_since_last_use
                    
                    # Track latest key usage for overall user activity
                    if not access_keys_info['latest_key_usage'] or last_used_date > access_keys_info['latest_key_usage']:
                        access_keys_info['latest_key_usage'] = last_used_date
                else:
                    key_info['DaysSinceLastUse'] = None
                
                access_keys_info['all_keys'].append(key_info)
                
                # Categorize old keys
                if key_age_days > days_threshold:
                    if not last_used_date:  # Never used
                        access_keys_info['old_unused_keys'].append(key_info)
                    else:
                        days_since_last_use = (today - last_used_date).days
                        if days_since_last_use > days_threshold:
                            access_keys_info['old_used_keys'].append(key_info)
        
        except Exception as e:
            self.logger.warning(f"Error analyzing access keys for {username}: {e}")
        
        return access_keys_info
    
    def analyze_single_user(self, iam, user, days_threshold, today):
        """Comprehensive analysis of a single user"""
        username = user['UserName']
        user_created_date = user['CreateDate']
        user_age_days = (today - user_created_date).days
        
        # Skip if filtered
        if self.filter_by_tags(iam, username):
            return None
        
        try:
            # Get user's password last used
            password_last_used = user.get('PasswordLastUsed')
            
            # Detailed access key analysis
            access_keys_info = {}
            if self.config['analyze_access_keys']:
                access_keys_info = self.analyze_access_keys_detailed(iam, username, today, days_threshold)
            else:
                # Basic access key info
                keys_response = self.make_api_call_with_retry(
                    iam.list_access_keys, UserName=username
                )
                access_keys = keys_response.get('AccessKeyMetadata', [])
                access_keys_info = {
                    'total_keys': len(access_keys),
                    'active_keys': len([k for k in access_keys if k['Status'] == 'Active']),
                    'all_keys': []
                }
            
            # Determine last activity
            last_activity = None
            activity_type = "Never"
            
            if password_last_used:
                last_activity = password_last_used
                activity_type = "Console"
            
            if access_keys_info.get('latest_key_usage'):
                if not last_activity or access_keys_info['latest_key_usage'] > last_activity:
                    last_activity = access_keys_info['latest_key_usage']
                    activity_type = "Access Key"
            
            # Calculate days since last activity
            days_since_activity = None
            if last_activity:
                days_since_activity = (today - last_activity).days
            
            # Check MFA status
            mfa_enabled = self.check_mfa_status(iam, username)
            
            # Policy analysis
            policies = []
            risky_policies = []
            if self.config['analyze_policies']:
                policies, risky_policies = self.analyze_policies(iam, username)
            
            user_data = {
                'Username': username,
                'Created': user_created_date,
                'AgeInDays': user_age_days,
                'LastActivity': last_activity,
                'ActivityType': activity_type,
                'DaysSinceActivity': days_since_activity,
                'AccessKeys': access_keys_info.get('total_keys', 0),
                'ActiveAccessKeys': access_keys_info.get('active_keys', 0),
                'AccessKeysDetail': access_keys_info,
                'MFAEnabled': mfa_enabled,
                'Policies': policies,
                'RiskyPolicies': risky_policies,
                'PolicyCount': len(policies),
                'is_inactive': not last_activity or days_since_activity > days_threshold
            }
            
            return user_data
            
        except Exception as e:
            self.logger.warning(f"Error processing user {username}: {e}")
            return None
    
    def get_all_iam_users(self, session, days_threshold):
        """Get all IAM users with comprehensive analysis using parallel processing"""
        iam = session.client('iam')
        
        # Check permissions
        if not self.check_required_permissions(iam):
            sys.exit(1)
        
        self.logger.info("Retrieving all IAM users...")
        
        try:
            # Get all users
            users = []
            paginator = iam.get_paginator('list_users')
            for page in paginator.paginate():
                users.extend(page['Users'])
            
            self.logger.info(f"Found {len(users)} IAM users")
            
            # Initialize data structures
            inactive_users = []
            active_users = []
            today = datetime.now(timezone.utc)
            
            # Process users in parallel
            self.logger.info("Processing users with comprehensive analysis...")
            with concurrent.futures.ThreadPoolExecutor(max_workers=self.config['max_workers']) as executor:
                user_futures = {
                    executor.submit(self.analyze_single_user, iam, user, days_threshold, today): user 
                    for user in users
                }
                
                for i, future in enumerate(concurrent.futures.as_completed(user_futures)):
                    if i > 0 and i % 10 == 0:
                        self.logger.info(f"Processed {i}/{len(users)} users...")
                    
                    user_data = future.result()
                    if user_data:
                        if user_data['is_inactive']:
                            inactive_users.append(user_data)
                        else:
                            active_users.append(user_data)
            
            self.logger.info(f"Successfully processed all {len(users)} users")
            
            # Collect access key statistics
            old_unused_keys = []
            old_used_keys = []
            
            if self.config['analyze_access_keys']:
                for user in inactive_users + active_users:
                    keys_detail = user.get('AccessKeysDetail', {})
                    for key in keys_detail.get('old_unused_keys', []):
                        old_unused_keys.append({
                            'Username': user['Username'],
                            **key
                        })
                    for key in keys_detail.get('old_used_keys', []):
                        old_used_keys.append({
                            'Username': user['Username'],
                            **key
                        })
            
            return {
                'inactive_users': inactive_users,
                'active_users': active_users,
                'total_users': len(users),
                'old_unused_keys': old_unused_keys,
                'old_used_keys': old_used_keys
            }
        
        except Exception as e:
            self.logger.error(f"Error retrieving IAM users: {e}")
            sys.exit(1)
    
    def display_comprehensive_analysis(self, results, days_threshold, account_id):
        """Display a comprehensive formatted report"""
        inactive_users = results['inactive_users']
        active_users = results['active_users']
        total_users = results['total_users']
        old_unused_keys = results.get('old_unused_keys', [])
        old_used_keys = results.get('old_used_keys', [])
        
        # Terminal width for better formatting
        terminal_width = 160
        
        # Helper functions
        def create_header(title):
            padding = (terminal_width - len(title)) // 2
            return "\n" + "=" * padding + " " + title + " " + "=" * padding
        
        def create_separator(char="-"):
            return char * terminal_width
        
        # Create report
        output = []
        output.append(create_separator("="))
        output.append("AWS IAM USER SECURITY ANALYSIS REPORT".center(terminal_width))
        output.append(f"Account ID: {account_id}".center(terminal_width))
        output.append(f"Inactivity Threshold: {days_threshold} days".center(terminal_width))
        output.append(f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}".center(terminal_width))
        output.append(create_separator("="))
        
        # Executive Summary
        output.append(create_header("EXECUTIVE SUMMARY"))
        output.append(f"\n📊 IAM Users Overview:")
        output.append(f"  Total IAM Users: {total_users}")
        output.append(f"    - Active Users (used within {days_threshold} days): {len(active_users)}")
        output.append(f"    - Inactive Users: {len(inactive_users)}")
        
        if self.config['analyze_access_keys']:
            output.append(f"\n🔑 Access Keys Analysis:")
            output.append(f"  Old Unused Keys (never used, >{days_threshold} days old): {len(old_unused_keys)}")
            output.append(f"  Old Used Keys (not used in >{days_threshold} days): {len(old_used_keys)}")
        
        # Security summary
        users_without_mfa = len([u for u in inactive_users + active_users if u.get('MFAEnabled') == False])
        users_with_risky_policies = len([u for u in inactive_users + active_users if u.get('RiskyPolicies')])
        
        output.append(f"\n🔒 Security Findings:")
        output.append(f"  Users without MFA: {users_without_mfa}")
        output.append(f"  Users with risky policies: {users_with_risky_policies}")
        
        # Calculate risk score
        total_issues = len(inactive_users) + len(old_unused_keys) + len(old_used_keys) + users_without_mfa
        risk_level = "LOW"
        if total_issues > 50:
            risk_level = "CRITICAL"
        elif total_issues > 20:
            risk_level = "HIGH"
        elif total_issues > 10:
            risk_level = "MEDIUM"
        
        output.append(f"\n⚠️  Overall Risk Level: {risk_level}")
        output.append(f"  Total security issues identified: {total_issues}")
        
        # ACCESS KEYS ANALYSIS SECTION
        if self.config['analyze_access_keys'] and (old_unused_keys or old_used_keys):
            output.append(create_header("ACCESS KEYS ANALYSIS"))
            
            # Old unused keys
            output.append(f"\n🔑 Old Unused Access Keys")
            output.append(f"Found {len(old_unused_keys)} access keys older than {days_threshold} days that have never been used\n")
            
            if old_unused_keys:
                sorted_keys = sorted(old_unused_keys, key=lambda x: x['AgeInDays'], reverse=True)
                
                for key in sorted_keys[:20]:  # Show top 20
                    output.append(f"Username: {key['Username']}")
                    output.append(f"  Access Key ID: {key['AccessKeyId']}")
                    output.append(f"  Created: {key['Created'].strftime('%Y-%m-%d %H:%M:%S')}")
                    output.append(f"  Age: {key['AgeInDays']} days")
                    output.append(f"  Status: {key['Status']}")
                    output.append("")
                
                if len(old_unused_keys) > 20:
                    output.append(f"... and {len(old_unused_keys) - 20} more (see Excel export for full list)")
            
            # Old used keys
            output.append(f"\n🔑 Old Used Access Keys (Not Recently Active)")
            output.append(f"Found {len(old_used_keys)} access keys not used in the last {days_threshold} days\n")
            
            if old_used_keys:
                sorted_keys = sorted(old_used_keys, key=lambda x: x.get('DaysSinceLastUse', 0), reverse=True)
                
                for key in sorted_keys[:20]:  # Show top 20
                    output.append(f"Username: {key['Username']}")
                    output.append(f"  Access Key ID: {key['AccessKeyId']}")
                    output.append(f"  Created: {key['Created'].strftime('%Y-%m-%d %H:%M:%S')}")
                    output.append(f"  Age: {key['AgeInDays']} days")
                    output.append(f"  Last Used: {key['LastUsed'].strftime('%Y-%m-%d %H:%M:%S') if key.get('LastUsed') else 'Never'}")
                    output.append(f"  Days Since Last Use: {key.get('DaysSinceLastUse', 'N/A')}")
                    output.append(f"  Last Used Service: {key.get('LastUsedService', 'N/A')}")
                    output.append(f"  Last Used Region: {key.get('LastUsedRegion', 'N/A')}")
                    output.append(f"  Status: {key['Status']}")
                    output.append("")
                
                if len(old_used_keys) > 20:
                    output.append(f"... and {len(old_used_keys) - 20} more (see Excel export for full list)")
        
        # INACTIVE USERS SECTION
        output.append(create_header("INACTIVE USERS ANALYSIS"))
        output.append(f"Found {len(inactive_users)} users with no activity in the last {days_threshold} days\n")
        
        if inactive_users:
            sorted_users = sorted(inactive_users, key=lambda x: x['DaysSinceActivity'] if x['DaysSinceActivity'] else 99999, reverse=True)
            
            for user in sorted_users[:15]:  # Show top 15
                output.append(f"👤 Username: {user['Username']}")
                output.append(f"  Created: {user['Created'].strftime('%Y-%m-%d %H:%M:%S')}")
                output.append(f"  Age: {user['AgeInDays']} days")
                output.append(f"  Last Activity: {user['LastActivity'].strftime('%Y-%m-%d %H:%M:%S') if user['LastActivity'] else 'Never'}")
                output.append(f"  Activity Type: {user['ActivityType']}")
                output.append(f"  Days Since Activity: {user['DaysSinceActivity'] if user['DaysSinceActivity'] else 'N/A'}")
                output.append(f"  Access Keys: {user['ActiveAccessKeys']}/{user['AccessKeys']} (Active/Total)")
                output.append(f"  MFA Enabled: {'✅ Yes' if user.get('MFAEnabled') else '❌ No' if user.get('MFAEnabled') == False else '❓ N/A'}")
                
                if self.config['analyze_access_keys']:
                    keys_detail = user.get('AccessKeysDetail', {})
                    if keys_detail.get('all_keys'):
                        output.append(f"  Access Keys Detail:")
                        for key in keys_detail['all_keys']:
                            output.append(f"    - {key['AccessKeyId']}: {key['Status']}, Age: {key['AgeInDays']} days, "
                                        f"Last Used: {key['LastUsed'].strftime('%Y-%m-%d') if key.get('LastUsed') else 'Never'}")
                
                if self.config['analyze_policies']:
                    output.append(f"  Policy Count: {user.get('PolicyCount', 0)}")
                    if user.get('RiskyPolicies'):
                        output.append(f"  ⚠️  Risky Policies: {', '.join(user['RiskyPolicies'])}")
                
                output.append("")
            
            if len(inactive_users) > 15:
                output.append(f"... and {len(inactive_users) - 15} more (see Excel export for full list)\n")
        else:
            output.append("✅ No inactive users found.")
        
        # MFA STATUS SECTION
        if users_without_mfa > 0:
            output.append(create_header("MFA STATUS - USERS WITHOUT MFA"))
            users_no_mfa = [u for u in inactive_users + active_users if u.get('MFAEnabled') == False]
            output.append(f"Found {len(users_no_mfa)} users without MFA enabled\n")
            
            for user in users_no_mfa[:20]:  # Show top 20
                status = "INACTIVE" if user in inactive_users else "Active"
                output.append(f"  - {user['Username']} ({status})")
            
            if len(users_no_mfa) > 20:
                output.append(f"  ... and {len(users_no_mfa) - 20} more")
        
        # POLICY ANALYSIS SECTION
        if self.config['analyze_policies']:
            output.append(create_header("RISKY POLICY ATTACHMENTS"))
            
            # High-risk users
            high_risk_users = [u for u in inactive_users + active_users if u.get('RiskyPolicies')]
            if high_risk_users:
                output.append(f"\n⚠️  Users with potentially risky policies ({len(high_risk_users)} found):")
                for user in high_risk_users[:20]:
                    policies_str = ", ".join(user.get('RiskyPolicies', []))
                    status = "INACTIVE" if user in inactive_users else "Active"
                    output.append(f"  - {user['Username']} ({status}): {policies_str}")
                
                if len(high_risk_users) > 20:
                    output.append(f"  ... and {len(high_risk_users) - 20} more")
            else:
                output.append("\n✅ No users found with risky policies.")
        
        # RECOMMENDATIONS SECTION
        output.append(create_separator("="))
        output.append("RECOMMENDATIONS AND ACTION ITEMS".center(terminal_width))
        output.append(create_separator("="))
        
        output.append(f"\n📋 Immediate Actions Required:")
        
        if old_unused_keys:
            output.append(f"\n1. ACCESS KEYS - Old Unused Keys ({len(old_unused_keys)} keys)")
            output.append("   • Delete unused access keys that have never been used")
            output.append("   • These keys pose a security risk with no benefit")
        
        if old_used_keys:
            output.append(f"\n2. ACCESS KEYS - Old Used Keys ({len(old_used_keys)} keys)")
            output.append("   • Rotate or delete access keys not used recently")
            output.append("   • Contact key owners to determine if still needed")
        
        if inactive_users:
            output.append(f"\n3. INACTIVE USERS ({len(inactive_users)} users)")
            output.append("   • Review and disable inactive user accounts")
            output.append("   • Delete users that are no longer needed")
            output.append("   • Document justification for keeping any inactive users")
        
        if users_without_mfa:
            output.append(f"\n4. MFA ENFORCEMENT ({users_without_mfa} users)")
            output.append("   • Require MFA for all users immediately")
            output.append("   • Consider using SCPs to enforce MFA")
            output.append("   • Disable console access for users without MFA")
        
        if users_with_risky_policies:
            output.append(f"\n5. PRIVILEGE REVIEW ({users_with_risky_policies} users)")
            output.append("   • Review administrative and full-access policies")
            output.append("   • Apply principle of least privilege")
            output.append("   • Replace broad permissions with specific ones")
        
        output.append("\n📚 Best Practices & Long-term Improvements:")
        output.append("   • Implement automated access key rotation (90-day maximum)")
        output.append("   • Set up AWS Config rules for continuous compliance monitoring")
        output.append("   • Enable CloudTrail for comprehensive IAM activity logging")
        output.append("   • Conduct quarterly IAM access reviews")
        output.append("   • Use IAM roles for applications instead of access keys")
        output.append("   • Implement AWS Organizations SCPs for guardrails")
        output.append("   • Enable IAM Access Analyzer for continuous monitoring")
        output.append("   • Document all service accounts and their purpose")
        
        output.append("\n🔍 Monitoring & Detection:")
        output.append("   • Set up CloudWatch alarms for:")
        output.append("     - Root account usage")
        output.append("     - Failed console login attempts")
        output.append("     - IAM policy changes")
        output.append("     - Access key creation/deletion")
        output.append("   • Enable AWS Security Hub for centralized findings")
        output.append("   • Consider AWS IAM Access Analyzer for external access detection")
        
        output.append("\n" + create_separator("="))
        output.append(f"Report complete. Total users analyzed: {total_users}".center(terminal_width))
        output.append(create_separator("="))
        
        # Print the report
        print("\n".join(output))
        
        return output
    
    def export_to_excel(self, results, filename_base, account_id, days_threshold):
        """Export comprehensive results to a single Excel file with multiple tabs"""
        if not EXCEL_AVAILABLE:
            self.logger.warning("Excel export not available. Falling back to CSV.")
            return self.export_to_csv(results, filename_base)
        
        excel_filename = f"{filename_base}.xlsx"
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # New styles for findings guide
        section_header_font = Font(bold=True, size=12, color="FFFFFF")
        section_header_fill = PatternFill(start_color="D9534F", end_color="D9534F", fill_type="solid")
        finding_title_font = Font(bold=True, size=10, color="000000")
        finding_title_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Helper function to style headers and auto-size columns
        def style_worksheet(ws, headers):
            # Style header row
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_style
            
            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = ws['A2']
        
        # TAB 1: Summary
        ws_summary = wb.create_sheet("Summary")
        
        # Basic summary data
        summary_data = [
            ["AWS IAM User Security Analysis Summary"],
            ["Account ID", account_id],
            ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ["Threshold (days)", days_threshold],
            [""],
            ["Category", "Count"],
            ["Total Users", results['total_users']],
            ["Active Users", len(results['active_users'])],
            ["Inactive Users", len(results['inactive_users'])],
        ]
        
        if self.config['analyze_access_keys']:
            summary_data.extend([
                ["Old Unused Access Keys", len(results.get('old_unused_keys', []))],
                ["Old Used Access Keys", len(results.get('old_used_keys', []))],
            ])
        
        users_without_mfa = len([u for u in results['inactive_users'] + results['active_users'] 
                                if u.get('MFAEnabled') == False])
        summary_data.append(["Users Without MFA", users_without_mfa])
        
        # Add summary data
        for row_data in summary_data:
            ws_summary.append(row_data)
        
        # Style summary section
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:B1')
        ws_summary['A6'].font = header_font
        ws_summary['B6'].font = header_font
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20
        
        # Add spacing before findings guide
        current_row = len(summary_data) + 3
        
        # Add Findings Guide section
        ws_summary.cell(row=current_row, column=1, value="SECURITY FINDINGS REFERENCE GUIDE")
        ws_summary.merge_cells(f'A{current_row}:B{current_row}')
        ws_summary.cell(row=current_row, column=1).font = section_header_font
        ws_summary.cell(row=current_row, column=1).fill = section_header_fill
        ws_summary.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        current_row += 2
        
        # Define findings guide content (user-focused only)
        findings_guide = [
            {
                "number": "1",
                "title": "Old Unused Access Keys",
                "risk": "🔴 HIGH - Credentials that could be compromised without detection",
                "criteria": "Access keys older than threshold that have never been used",
                "recommendation": "Delete immediately"
            },
            {
                "number": "2",
                "title": "Old Used Access Keys",
                "risk": "🟡 MEDIUM-HIGH - Potentially compromised or forgotten credentials",
                "criteria": "Access keys not used within threshold period",
                "recommendation": "Rotate or delete after confirming they're no longer needed"
            },
            {
                "number": "3",
                "title": "Inactive Users",
                "risk": "⚠️ MEDIUM - Abandoned accounts that could be compromised",
                "criteria": "No console or programmatic access within threshold period",
                "recommendation": "Disable or delete unused accounts"
            },
            {
                "number": "4",
                "title": "Users Without MFA",
                "risk": "🔴 HIGH - Vulnerable to credential theft and account takeover",
                "criteria": "IAM users without MFA devices configured",
                "recommendation": "Enforce MFA immediately"
            },
            {
                "number": "5",
                "title": "Risky Policy Attachments",
                "risk": "🔴 HIGH - Excessive privileges that violate least privilege principle",
                "criteria": "Policies containing 'Admin', 'Full', '*', or 'PowerUser' in name",
                "recommendation": "Review and apply least privilege principle"
            }
        ]
        
        # Add each finding to the guide
        for finding in findings_guide:
            # Finding title row
            ws_summary.cell(row=current_row, column=1, value=f"{finding['number']}. {finding['title']}")
            ws_summary.merge_cells(f'A{current_row}:B{current_row}')
            ws_summary.cell(row=current_row, column=1).font = finding_title_font
            ws_summary.cell(row=current_row, column=1).fill = finding_title_fill
            ws_summary.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1
            
            # Risk row
            ws_summary.cell(row=current_row, column=1, value="Risk:")
            ws_summary.cell(row=current_row, column=2, value=finding['risk'])
            ws_summary.cell(row=current_row, column=1).font = Font(bold=True, size=9)
            ws_summary.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
            current_row += 1
            
            # Criteria row
            ws_summary.cell(row=current_row, column=1, value="Criteria:")
            ws_summary.cell(row=current_row, column=2, value=finding['criteria'])
            ws_summary.cell(row=current_row, column=1).font = Font(bold=True, size=9)
            ws_summary.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
            current_row += 1
            
            # Recommendation row
            ws_summary.cell(row=current_row, column=1, value="Recommendation:")
            ws_summary.cell(row=current_row, column=2, value=finding['recommendation'])
            ws_summary.cell(row=current_row, column=1).font = Font(bold=True, size=9)
            ws_summary.cell(row=current_row, column=2).alignment = Alignment(wrap_text=True)
            ws_summary.row_dimensions[current_row].height = 30
            current_row += 2
        
        # Adjust column B width for better text wrapping
        ws_summary.column_dimensions['B'].width = 80
        
        # TAB 2: All Users
        ws_users = wb.create_sheet("All Users")
        user_headers = [
            'Username', 'Status', 'Created', 'Age (Days)', 'Last Activity', 
            'Activity Type', 'Days Since Activity', 'Total Keys', 'Active Keys',
            'Old Unused Keys', 'Old Used Keys', 'MFA Enabled', 'Policy Count', 
            'Risky Policies', 'All Policies', 'Managed Policies', 'Inline Policies'
        ]
        ws_users.append(user_headers)
        
        for user in results['inactive_users'] + results['active_users']:
            managed_policies = []
            inline_policies = []
            
            for policy in user.get('Policies', []):
                if policy.get('Type') == 'Managed':
                    managed_policies.append(policy.get('PolicyName', ''))
                elif policy.get('Type') == 'Inline':
                    inline_policies.append(policy.get('PolicyName', ''))
            
            keys_detail = user.get('AccessKeysDetail', {})
            
            row = [
                user['Username'],
                'Inactive' if user in results['inactive_users'] else 'Active',
                user['Created'].strftime('%Y-%m-%d %H:%M:%S') if user['Created'] else '',
                user['AgeInDays'],
                user['LastActivity'].strftime('%Y-%m-%d %H:%M:%S') if user['LastActivity'] else 'Never',
                user['ActivityType'],
                user['DaysSinceActivity'] if user['DaysSinceActivity'] is not None else '',
                user['AccessKeys'],
                user['ActiveAccessKeys'],
                len(keys_detail.get('old_unused_keys', [])),
                len(keys_detail.get('old_used_keys', [])),
                'Yes' if user.get('MFAEnabled') else 'No' if user.get('MFAEnabled') == False else 'N/A',
                user.get('PolicyCount', 0),
                '; '.join(user.get('RiskyPolicies', [])),
                '; '.join([f"{p.get('PolicyName', '')} ({p.get('Type', '')})" for p in user.get('Policies', [])]),
                '; '.join(managed_policies),
                '; '.join(inline_policies)
            ]
            ws_users.append(row)
        
        style_worksheet(ws_users, user_headers)
        
        # Conditional formatting for inactive users
        for row in range(2, ws_users.max_row + 1):
            status_cell = ws_users[f'B{row}']
            mfa_cell = ws_users[f'L{row}']
            
            if status_cell.value == 'Inactive':
                for col in range(1, len(user_headers) + 1):
                    ws_users.cell(row, col).fill = warning_fill
            
            if mfa_cell.value == 'No':
                mfa_cell.font = Font(color="FF0000", bold=True)
        
        # TAB 3: Old Unused Access Keys
        if self.config['analyze_access_keys'] and results.get('old_unused_keys'):
            ws_unused_keys = wb.create_sheet("Old Unused Keys")
            unused_key_headers = ['Username', 'Access Key ID', 'Status', 'Created', 'Age (Days)']
            ws_unused_keys.append(unused_key_headers)
            
            for key in results['old_unused_keys']:
                row = [
                    key['Username'],
                    key['AccessKeyId'],
                    key['Status'],
                    key['Created'].strftime('%Y-%m-%d %H:%M:%S'),
                    key['AgeInDays']
                ]
                ws_unused_keys.append(row)
            
            style_worksheet(ws_unused_keys, unused_key_headers)
            
            # Highlight all rows as these are security issues
            for row in range(2, ws_unused_keys.max_row + 1):
                for col in range(1, len(unused_key_headers) + 1):
                    ws_unused_keys.cell(row, col).fill = warning_fill
        
        # TAB 4: Old Used Access Keys
        if self.config['analyze_access_keys'] and results.get('old_used_keys'):
            ws_used_keys = wb.create_sheet("Old Used Keys")
            used_key_headers = [
                'Username', 'Access Key ID', 'Status', 'Created', 'Age (Days)',
                'Last Used', 'Days Since Last Use', 'Last Used Service', 'Last Used Region'
            ]
            ws_used_keys.append(used_key_headers)
            
            for key in results['old_used_keys']:
                row = [
                    key['Username'],
                    key['AccessKeyId'],
                    key['Status'],
                    key['Created'].strftime('%Y-%m-%d %H:%M:%S'),
                    key['AgeInDays'],
                    key['LastUsed'].strftime('%Y-%m-%d %H:%M:%S') if key.get('LastUsed') else 'Never',
                    key.get('DaysSinceLastUse', ''),
                    key.get('LastUsedService', 'N/A'),
                    key.get('LastUsedRegion', 'N/A')
                ]
                ws_used_keys.append(row)
            
            style_worksheet(ws_used_keys, used_key_headers)
            
            # Highlight all rows as these are security issues
            for row in range(2, ws_used_keys.max_row + 1):
                for col in range(1, len(used_key_headers) + 1):
                    ws_used_keys.cell(row, col).fill = warning_fill
        
        # TAB 5: Users Without MFA
        users_no_mfa = [u for u in results['inactive_users'] + results['active_users'] 
                       if u.get('MFAEnabled') == False]
        if users_no_mfa:
            ws_no_mfa = wb.create_sheet("Users Without MFA")
            mfa_headers = ['Username', 'Status', 'Last Activity', 'Days Since Activity', 'Active Keys']
            ws_no_mfa.append(mfa_headers)
            
            for user in users_no_mfa:
                row = [
                    user['Username'],
                    'Inactive' if user in results['inactive_users'] else 'Active',
                    user['LastActivity'].strftime('%Y-%m-%d %H:%M:%S') if user['LastActivity'] else 'Never',
                    user['DaysSinceActivity'] if user['DaysSinceActivity'] is not None else 'N/A',
                    user['ActiveAccessKeys']
                ]
                ws_no_mfa.append(row)
            
            style_worksheet(ws_no_mfa, mfa_headers)
            
            # Highlight all rows as these are security issues
            for row in range(2, ws_no_mfa.max_row + 1):
                for col in range(1, len(mfa_headers) + 1):
                    ws_no_mfa.cell(row, col).fill = warning_fill
        
        # TAB 6: Risky Policies
        if self.config['analyze_policies']:
            high_risk_users = [u for u in results['inactive_users'] + results['active_users'] 
                             if u.get('RiskyPolicies')]
            
            if high_risk_users:
                ws_risky = wb.create_sheet("Risky Policies")
                risky_headers = ['Username', 'Status', 'Risky Policies']
                ws_risky.append(risky_headers)
                
                for user in high_risk_users:
                    row = [
                        user['Username'],
                        'Inactive' if user in results['inactive_users'] else 'Active',
                        '; '.join(user.get('RiskyPolicies', []))
                    ]
                    ws_risky.append(row)
                
                style_worksheet(ws_risky, risky_headers)
                
                # Highlight all rows as these are security concerns
                for row in range(2, ws_risky.max_row + 1):
                    for col in range(1, len(risky_headers) + 1):
                        ws_risky.cell(row, col).fill = warning_fill
        
        # Save workbook
        wb.save(excel_filename)
        self.logger.info(f"Excel file created: {excel_filename}")
        
        return [excel_filename]
    
    def export_to_csv(self, results, filename_base):
        """Export comprehensive results to CSV format (fallback if Excel not available)"""
        files_created = []
        
        # Export users with all details
        users_filename = f"{filename_base}_users.csv"
        with open(users_filename, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = [
                'Username', 'Status', 'Created', 'AgeInDays', 'LastActivity', 
                'ActivityType', 'DaysSinceActivity', 'TotalAccessKeys', 'ActiveAccessKeys',
                'OldUnusedKeys', 'OldUsedKeys', 'MFAEnabled', 'PolicyCount', 
                'RiskyPolicies', 'AllPolicies', 'ManagedPolicies', 'InlinePolicies'
            ]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            
            # Process all users (inactive + active)
            for user in results['inactive_users'] + results['active_users']:
                managed_policies = []
                inline_policies = []
                
                for policy in user.get('Policies', []):
                    if policy.get('Type') == 'Managed':
                        managed_policies.append(policy.get('PolicyName', ''))
                    elif policy.get('Type') == 'Inline':
                        inline_policies.append(policy.get('PolicyName', ''))
                
                keys_detail = user.get('AccessKeysDetail', {})
                
                row = {
                    'Username': user['Username'],
                    'Status': 'Inactive' if user in results['inactive_users'] else 'Active',
                    'Created': user['Created'].strftime('%Y-%m-%d %H:%M:%S') if user['Created'] else '',
                    'AgeInDays': user['AgeInDays'],
                    'LastActivity': user['LastActivity'].strftime('%Y-%m-%d %H:%M:%S') if user['LastActivity'] else 'Never',
                    'ActivityType': user['ActivityType'],
                    'DaysSinceActivity': user['DaysSinceActivity'] if user['DaysSinceActivity'] is not None else '',
                    'TotalAccessKeys': user['AccessKeys'],
                    'ActiveAccessKeys': user['ActiveAccessKeys'],
                    'OldUnusedKeys': len(keys_detail.get('old_unused_keys', [])),
                    'OldUsedKeys': len(keys_detail.get('old_used_keys', [])),
                    'MFAEnabled': user.get('MFAEnabled', ''),
                    'PolicyCount': user.get('PolicyCount', 0),
                    'RiskyPolicies': '; '.join(user.get('RiskyPolicies', [])),
                    'AllPolicies': '; '.join([f"{p.get('PolicyName', '')} ({p.get('Type', '')})" for p in user.get('Policies', [])]),
                    'ManagedPolicies': '; '.join(managed_policies),
                    'InlinePolicies': '; '.join(inline_policies)
                }
                writer.writerow(row)
        
        files_created.append(users_filename)
        
        # Export access keys details if analyzed
        if self.config['analyze_access_keys']:
            # Old unused keys
            if results.get('old_unused_keys'):
                unused_keys_filename = f"{filename_base}_old_unused_keys.csv"
                with open(unused_keys_filename, 'w', newline='', encoding='utf-8') as csvfile:
                    fieldnames = ['Username', 'AccessKeyId', 'Status', 'Created', 'AgeInDays']
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for key in results['old_unused_keys']:
                        row = {
                            'Username': key['Username'],
                            'AccessKeyId': key['AccessKeyId'],
                            'Status': key['Status'],
                            'Created': key['Created'].strftime('%Y-%m-%d %H:%M:%S'),
                            'AgeInDays': key['AgeInDays']
                        }
                        writer.writerow(row)
                
                files_created.append(unused_keys_filename)
            
            # Old used keys
            if results.get('old_used_keys'):
                used_keys_filename = f"{filename_base}_old_used_keys.csv"
                with open(used_keys_filename, 'w', newline='', encoding='utf-8') as csvfile:
                    fieldnames = [
                        'Username', 'AccessKeyId', 'Status', 'Created', 'AgeInDays',
                        'LastUsed', 'DaysSinceLastUse', 'LastUsedService', 'LastUsedRegion'
                    ]
                    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                    writer.writeheader()
                    
                    for key in results['old_used_keys']:
                        row = {
                            'Username': key['Username'],
                            'AccessKeyId': key['AccessKeyId'],
                            'Status': key['Status'],
                            'Created': key['Created'].strftime('%Y-%m-%d %H:%M:%S'),
                            'AgeInDays': key['AgeInDays'],
                            'LastUsed': key['LastUsed'].strftime('%Y-%m-%d %H:%M:%S') if key.get('LastUsed') else 'Never',
                            'DaysSinceLastUse': key.get('DaysSinceLastUse', ''),
                            'LastUsedService': key.get('LastUsedService', 'N/A'),
                            'LastUsedRegion': key.get('LastUsedRegion', 'N/A')
                        }
                        writer.writerow(row)
                
                files_created.append(used_keys_filename)
        
        return files_created
    
    def save_results_to_files(self, results, account_id, profile_name, days_threshold):
        """Save comprehensive results to multiple file formats"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename_base = f"iam_user_security_report_{account_id}_{timestamp}"
        
        files_created = []
        
        # Create text report
        if 'txt' in self.config['output_formats']:
            text_filename = f"{filename_base}.txt"
            report_lines = self.display_comprehensive_analysis(results, days_threshold, account_id)
            
            with open(text_filename, 'w', encoding='utf-8') as text_file:
                text_file.write("\n".join(report_lines))
            files_created.append(text_filename)
        
        # Create Excel or CSV exports
        if 'excel' in self.config['output_formats']:
            excel_files = self.export_to_excel(results, filename_base, account_id, days_threshold)
            files_created.extend(excel_files)
        elif 'csv' in self.config['output_formats']:
            csv_files = self.export_to_csv(results, filename_base)
            files_created.extend(csv_files)
        
        return files_created

def main():
    try:
        print("\n" + "="*80)
        print("AWS IAM USER SECURITY ANALYZER".center(80))
        print("="*80)
        print("\n🔍 This script provides:")
        print("  ✓ Complete IAM user activity analysis")
        print("  ✓ Detailed access key analysis (age, usage, last used service/region)")
        print("  ✓ Policy attachment analysis with risk assessment")
        print("  ✓ MFA status checking for all users")
        print("  ✓ Parallel processing for improved performance")
        print("  ✓ Excel export with multiple tabs (or CSV fallback)")
        print("  ✓ Configurable filtering and exclusions")
        print("  ✓ Comprehensive security recommendations")
        print("  ✓ Risk scoring and prioritization")
        
        if not EXCEL_AVAILABLE:
            print("\n⚠️  Note: openpyxl not installed. Install with 'pip install openpyxl' for Excel export.")
        
        # Initialize analyzer
        analyzer = IAMUserSecurityAnalyzer()
        
        # Get user configuration
        analyzer.get_user_configuration()
        
        # Get AWS profile and account ID
        session, account_id, profile_name = analyzer.get_profile_and_account_info()
        
        if not analyzer.config['dry_run']:
            proceed = input("\nProceed with user security analysis? (y/n): ").lower()
            if proceed not in ['y', 'yes']:
                print("Analysis cancelled.")
                sys.exit(0)
        
        print(f"\n🔍 Starting IAM user security analysis...")
        print(f"   Threshold: {analyzer.config['days_threshold']} days")
        print(f"   Workers: {analyzer.config['max_workers']} parallel threads")
        start_time = time.time()
        
        # Get all IAM users with comprehensive analysis
        results = analyzer.get_all_iam_users(session, analyzer.config['days_threshold'])
        
        # Save results to files
        files_created = analyzer.save_results_to_files(
            results, account_id, profile_name, analyzer.config['days_threshold']
        )
        
        end_time = time.time()
        analysis_duration = end_time - start_time
        
        print(f"\n✅ Analysis completed in {analysis_duration:.2f} seconds")
        print(f"\n📄 Files created:")
        for file_path in files_created:
            file_size = os.path.getsize(file_path)
            print(f"   • {file_path} ({file_size:,} bytes)")
        
        # Summary statistics
        total_inactive = len(results['inactive_users'])
        total_key_issues = len(results.get('old_unused_keys', [])) + len(results.get('old_used_keys', []))
        
        print(f"\n📊 Analysis Summary:")
        print(f"   • Total users analyzed: {results['total_users']}")
        print(f"   • Inactive users: {total_inactive}")
        
        if analyzer.config['analyze_access_keys']:
            print(f"   • Access key issues: {total_key_issues}")
        
        # Security alerts
        if total_inactive > 0:
            print(f"\n⚠️  Security Alert: Found {total_inactive} inactive IAM users requiring review!")
        
        if total_key_issues > 0:
            print(f"⚠️  Access Key Alert: Found {total_key_issues} old or unused access keys!")
        
        users_without_mfa = [u for u in results['inactive_users'] + results['active_users'] 
                           if u.get('MFAEnabled') == False]
        if users_without_mfa:
            print(f"🔒 MFA Alert: {len(users_without_mfa)} users without MFA enabled")
        
        if analyzer.config['analyze_policies']:
            risky_users = [u for u in results['inactive_users'] + results['active_users'] 
                         if u.get('RiskyPolicies')]
            if risky_users:
                print(f"🚨 Policy Alert: {len(risky_users)} users with potentially risky policies")
        
        if total_inactive == 0 and total_key_issues == 0 and not users_without_mfa:
            print(f"\n✅ Excellent! No critical security issues found.")
        
        print(f"\n📋 Review the generated reports for detailed findings and recommendations.")
        
    except KeyboardInterrupt:
        print("\n\n⏹️  Operation cancelled by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        logging.error(f"Unexpected error: {e}", exc_info=True)
        sys.exit(1)

if __name__ == "__main__":
    main()
